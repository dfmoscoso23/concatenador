#Control de Flex

from tkinter import*
from tkinter import ttk
from tkinter import messagebox as mb
import openpyxl as op
from datetime import date, timedelta, datetime
from fpdf import FPDF

#valores:
flexcaba=250
flexgba1=400
flexgba2=500

mercadocaba=309.99
mercadogba1=479.99
mercadogba2=659.99

today = date.today()

#archivos
wb = op.load_workbook('flexcontrol.xlsx')
#diccionarios
operacionespendientes = []
dic_tipodeoperacion = {}
dic_destino = {}
dic_factura={}
dic_bonificacion = {}
dic_vendedor = {}
dic_vgenerales={'CABA':0,'GBA 1':0,'GBA 2':0}
dic_precios={'CABA':mercadocaba,'GBA 1':mercadogba1,'GBA 2':mercadogba2}
pen = 0
#Definiciones
def calculo(u, sheet, destino, vendedorc):
	lis_dest = []
	lis_bon=[]
	lis_ent=[]
	lis_prop=[]
	for x in range(1,u):
		if sheet.cell(column=5, row=x).value==vendedorc:
			xdestino=sheet.cell(column=3, row=x).value
			xbonificado = sheet.cell(column=4, row=x).value
			xtipo = sheet.cell(column=2, row=x).value
			if xdestino == destino:
				lis_dest.append('x')
				if xtipo == 2:
					lis_prop.append('x')
				else:
					if xbonificado==True:
						lis_bon.append('x')
					else:
						lis_ent.append('x')
	global drecibido
	drecibido += float((int(len(lis_bon)) * round(dic_precios[destino]/2))+(int(len(lis_ent))*dic_precios[destino]))
	dic_vgenerales[destino]+=len(lis_dest)
def detallador(u,sheet,destino,vendedorc):
	lis_dest=[]
	for x in range(1,u):
		if sheet.cell(column=5, row=x).value==vendedorc:
			xdestino=sheet.cell(column=3, row=x).value
			if xdestino == destino:
				lis_dest.append('x')
	return len(lis_dest)			
def calcular(fecha, vendedor):
	ini = fecha
	fin = ini + timedelta(days=-7)
	global drecibido
	drecibido = int(0)

	try:
		while ini >= fin:
			sheet = wb[str(ini)]
			u = (sheet.max_row + 1)
			#calculocaba(u, sheet)
			#calculogba1(u, sheet)
			#calculogba2(u, sheet)
			calculo(u,sheet,'CABA',vendedor)
			calculo(u,sheet,'GBA 1',vendedor)
			calculo(u,sheet,'GBA 2',vendedor)
			ini = ini + timedelta(days=-1)
	except KeyError:
		pass
	pagaraflex = float((dic_vgenerales['CABA']*flexcaba)+(dic_vgenerales['GBA 1']*flexgba1)+(dic_vgenerales['GBA 2']*flexgba2))
	totalapagar = float(pagaraflex-drecibido)
	global rframe
	bell2 = Label(rframe, text=dic_vgenerales['CABA'])
	bell2.grid(column=2, row=2)
	bell = Label(rframe, text=dic_vgenerales['GBA 1'])
	bell.grid(column=2, row=3)
	bell3 = Label(rframe, text=dic_vgenerales['GBA 2'])
	bell3.grid(column=2, row=4)
	bell4 = Label(rframe, text='${:,.2f}'.format(pagaraflex))
	bell4.grid(column=2, row=5)
	bell5 = Label(rframe, text='${:,.2f}'.format(drecibido))
	bell5.grid(column=2, row=6)
	bell6 = Label(rframe, text= '${:,.2f}'.format(totalapagar))
	bell6.grid(column=2, row=7)

def precalcular():
	global ingre_fecha
	fecharaw = ingre_fecha.get()
	year, month, day = map(int, fecharaw.split('-'))
	fecha = date(year, month, day)
	global check_losada_var
	global check_vallarta_var
	vendlosada = check_losada_var.get()
	vendvallarta = check_vallarta_var.get()
	if vendlosada == True and vendvallarta==False:
		vendedorc='Losada'
	elif vendlosada == False and vendvallarta==True:
		vendedorc='Vallarta'
	else:
		vendedorc='nada'
	calcular(fecha,vendedorc)
	dic_vgenerales['CABA']=0
	dic_vgenerales['GBA 1']=0
	dic_vgenerales['GBA 2']=0
def precalcular2():
	global ingre_fecha
	fecharaw = ingre_fecha.get()
	year, month, day = map(int, fecharaw.split('-'))
	fecha = date(year, month, day)
	global check_losada_var
	global check_vallarta_var
	vendlosada = check_losada_var.get()
	vendvallarta = check_vallarta_var.get()
	if vendlosada == True and vendvallarta==False:
		vendedorc='Losada'
	elif vendlosada == False and vendvallarta==True:
		vendedorc='Vallarta'
	else:
		vendedorc='nada'
	detalles(fecha,vendedorc)	

def detalles(fecha,vendedorc):
	global calc
	dframe = Frame(calc)
	dframe.grid(column=0, row=3)
	inis = []
	ini = fecha
	fin = ini + timedelta(days=-7)
	global drecibido
	drecibido = int(0)
	try:
		while ini >= fin:
			inis.append('x')
			l = len(inis)
			sheet = wb[str(ini)]
			u = (sheet.max_row + 1)
			detaf = Label(dframe, text=str(ini))
			detaf.grid(column=(l+1), row=1)
			deta = Label(dframe, text=detallador(u,sheet,'CABA',vendedorc))
			deta.grid(column=(l+1), row=2)
			deta2 = Label(dframe, text=detallador(u,sheet,'GBA 1',vendedorc))
			deta2.grid(column=(l+1), row=3)
			deta3 = Label(dframe, text=detallador(u,sheet,'GBA 2',vendedorc))
			deta3.grid(column=(l+1), row=4)
			ini = ini + timedelta(days=-1)
	except KeyError:
		pass
	deta4 = Label(dframe, text="CABA")
	deta4.grid(column=1, row=2)
	deta5 = Label(dframe, text="GBA1")
	deta5.grid(column=1, row=3)
	deta6 = Label(dframe, text="GBA2")
	deta6.grid(column=1, row=4)
def exportarinforme():
	inis=[]
	global ingre_fecha
	fecharaw = ingre_fecha.get()
	year, month, day = map(int, fecharaw.split('-'))
	fecha = date(year, month, day)
	ini = fecha
	fin = ini + timedelta(days=-7)
	global check_losada_var
	global check_vallarta_var
	vendlosada = check_losada_var.get()
	vendvallarta = check_vallarta_var.get()
	if vendlosada == True and vendvallarta==False:
		vendedorc='Losada'
	elif vendlosada == False and vendvallarta==True:
		vendedorc='Vallarta'
	else:
		vendedorc='nada'
	pdf = FPDF(orientation='P', format ='A4')
	pdf.add_page()
	pdf.set_font("Arial",'B', size=15)
	pdf.text((5),(5), txt="Librería Losada - Informe de envíos por Flex")
	global drecibido
	drecibido = int(0)
	rx=25
	try:
		while ini >= fin:
			pdf.set_font("Arial", size=10)
			sheet = wb[str(ini)]
			u = (sheet.max_row + 1)
			calculo(u,sheet,'CABA',vendedorc)
			calculo(u,sheet,'GBA 1',vendedorc)
			calculo(u,sheet,'GBA 2',vendedorc)
			inis.append('x')
			l = len(inis)
			sheet = wb[str(ini)]
			u = (sheet.max_row + 1)
			pdf.text(rx,80,txt=str(ini))
			pdf.text(rx+5,85,txt=str(detallador(u,sheet,'CABA',vendedorc)))
			pdf.text(rx+5,90,txt=str(detallador(u,sheet,'GBA 1',vendedorc)))
			pdf.text(rx+5,95,txt=str(detallador(u,sheet,'GBA 2',vendedorc)))		
			rx+=20
			ini = ini + timedelta(days=-1)
	except KeyError:
		pass
	pagaraflex = float((dic_vgenerales['CABA']*flexcaba)+(dic_vgenerales['GBA 1']*flexgba1)+(dic_vgenerales['GBA 2']*flexgba2))
	totalapagar = float(pagaraflex-drecibido)	
	pdf.set_font("Arial",'B', size=15)	
	pdf.text(5,15,txt="Vendedor:")
	pdf.text(45,15, txt=str(vendedorc))
	pdf.text(5,25,txt="Cantidad de envíos en CABA:")
	pdf.text(85,25, txt=str(dic_vgenerales['CABA']))
	pdf.text(5,30,txt="Cantidad de envíos en GBA 1:")
	pdf.text(85,30, txt=str(dic_vgenerales['GBA 1']))
	pdf.text(5,35,txt="Cantidad de envíos en GBA 2:")
	pdf.text(85,35, txt=str(dic_vgenerales['GBA 2']))
	pdf.text(5,40,txt="Monto a pagar a Flex:")
	pdf.text(65,40, txt=str('${:,.2f}'.format(pagaraflex)))
	pdf.text(5,45,txt="Dinero recibido:")
	pdf.text(65,45, txt=str('${:,.2f}'.format(drecibido)))
	pdf.text(5,50,txt="Total:")
	pdf.text(65,50, txt=str('${:,.2f}'.format(totalapagar)))
	pdf.text(5,75, txt=">>>>>>>>>>>>>>>>>DETALLE>>>>>>>>>>>>>>>>>>>>>")
	pdf.text(5,85,txt="CABA:")
	pdf.text(5,90,txt="GBA 1:")
	pdf.text(5,95,txt="GBA 2:")	
	ini=fecha
	pdf.add_page('L')
	pdf.text(5,5, txt=">>>>>>>>>>>>>>>>>OPERACIÓN ML vs FACTURA>>>>>>>>>>>>>>>>>>>>>")
	sx=5
	try:
		while ini >= fin:
			sheet = wb[str(ini)]
			pdf.set_font("Arial", size=7)
			pdf.text(sx,10,txt=str(ini))
			lar=15
			for ope in range(1,u):
				if ope == 1:
					pdf.text(sx,lar,txt=str(sheet.cell(column=1, row=ope).value)+">"+str(sheet.cell(column=6, row=ope).value)+">"+str(sheet.cell(column=3, row=ope).value))
					lar+=5
				elif sheet.cell(column=5, row=ope).value==vendedorc:
					pdf.text(sx,lar,txt=str(sheet.cell(column=1, row=ope).value)+">"+str(sheet.cell(column=6, row=ope).value)+">"+str(sheet.cell(column=3, row=ope).value))
					lar+=5		
			sx+=40
			rx+=20
			ini = ini + timedelta(days=-1)
	except KeyError:
		pass
	try:			
		pdf.output('informeenviosflexsemana'+vendedorc+str(fecha)+'.pdf','F')
		mb.showinfo("Aviso","Exportado con éxito")
	except PermissionError:
		mb.showerror("ERROR","PDF abierto")

	dic_vgenerales['CABA']=0
	dic_vgenerales['GBA 1']=0
	dic_vgenerales['GBA 2']=0
def ventanacalcular():
	today = date.today()
	global calc
	calc = Toplevel(raiz)
	aframe = Frame(calc)
	aframe.grid(column=0,row=0)
	global ingre_fecha
	ingre_fecha_label = Label(aframe, text="Último día de la semana: (AAAA-MM-DD):")
	ingre_fecha_label.grid(column=1, row=1)
	ingre_fecha = Entry(aframe, width=10)
	ingre_fecha.insert(END, today)
	ingre_fecha.grid(column=2, row=1)
	global check_vallarta_var
	check_vallarta_var = BooleanVar()
	check_vallarta=Checkbutton(aframe,text="Vallarta",var=check_vallarta_var)
	check_vallarta.grid(column=0, row=2)
	global check_losada_var
	check_losada_var = BooleanVar()
	check_losada=Checkbutton(aframe,text="Losada",var=check_losada_var)
	check_losada.grid(column=0, row=3)
	global rframe
	rframe = Frame(calc)
	rframe.grid(column=0,row=1)
	resul_label =Label(rframe, text="Resultados:")
	resul_label.grid(column=1, row=1)
	resul_caba_label =Label(rframe, text="Cantidad de envíos en CABA:")
	resul_caba_label.grid(column=1, row=2)
	resul_gba1_label =Label(rframe, text="Cantidad de envíos en GBA 1:")
	resul_gba1_label.grid(column=1, row=3)
	resul_gba2_label =Label(rframe, text="Cantidad de envíos en GBA 2:")
	resul_gba2_label.grid(column=1, row=4)
	resul_paga_label =Label(rframe, text="Cantidad a pagar:")
	resul_paga_label.grid(column=1, row=5)
	resul_paga2_label =Label(rframe, text="Cantidad de dinero recibido:")
	resul_paga2_label.grid(column=1, row=6)
	resul_paga3_label =Label(rframe, text="Saldo a pagar:")
	resul_paga3_label.grid(column=1, row=7)

	boton_calcular = Button(aframe, text="Calcular", command=precalcular)
	boton_calcular.grid(column=1, row=2)
	boton_detalles = Button(aframe, text="Detalles", command=precalcular2)
	boton_detalles.grid(column=2, row=2)
	boton_exportar = Button(aframe, text="Exportar Informe", command=exportarinforme)
	boton_exportar.grid(column=3, row=2)

def botonguardar():
	numoperacion = numero_operacion.get()
	if numoperacion not in operacionespendientes:
		operacionespendientes.append(numoperacion)
	tipooperacion = opcion.get()
	dic_tipodeoperacion[numoperacion] = tipooperacion
	destino2 = lugar.get()
	dic_destino[numoperacion] = destino2
	bonif = chk_bonificado_state.get()
	if bonif == 0:
		bonif = False
	dic_bonificacion[numoperacion] = bonif
	dic_vendedor[numoperacion]=vendedor.get()
	dic_factura[numoperacion]=numero_factura.get()
	hojas = wb.sheetnames
	if str(today) not in hojas:
		wb.create_sheet(str(today))
		sheet = wb[str(today)]
		sheet['A1'].value = 'Número de operación'
		sheet['B1'].value = 'Tipo'
		sheet['C1'].value = 'Destino'
		sheet['D1'].value = 'Bonificado'
		sheet['E1'].value = 'Vendedor'
		sheet['F1'].value = 'Factura'
	else:
		sheet = wb[str(today)]

	u = (sheet.max_row + 1)
	guar = False
	for x in range(1, u):
		xrow = sheet.cell(column=1, row=x).value
		if str(xrow) == str(numoperacion):
			sheet.cell(column=2, row=x).value = dic_tipodeoperacion[numoperacion]
			sheet.cell(column=3, row=x).value = dic_destino[numoperacion]
			sheet.cell(column=4, row=x).value = dic_bonificacion[numoperacion]
			sheet.cell(column=6, row=x).value = dic_factura[numoperacion]
			if dic_vendedor[numoperacion] == 1:
				sheet.cell(column=5, row=x).value = "Vallarta"
			else:
				sheet.cell(column=5, row=x).value = "Losada"	
			guar = True
			labell = Label(frame, text="Se ha modificado con exito")
			labell.grid(column=3, row=3)
			prenum = Label(preframe, text= numoperacion)
			prenum.grid(column=1, row=x)
			predes = Label(preframe, text= dic_destino[numoperacion])
			predes.grid(column=2, row=x)
			prebon = Label(preframe, text= dic_bonificacion[numoperacion])
			prebon.grid(column=3, row=x)
			if dic_vendedor[numoperacion] == 1: 
				preven = Label(preframe, text= "Vallarta")
				preven.grid(column=4, row=x)
			else:
				preven = Label(preframe, text= "Losada")
				preven.grid(column=4, row=x)
			prefac = Label(preframe, text= dic_factura[numoperacion])
			prefac.grid(column=5, row=x)		
	if guar == False:
		sheet.cell(column=1, row=u).value = numoperacion
		sheet.cell(column=2, row=u).value = dic_tipodeoperacion[numoperacion]
		sheet.cell(column=3, row=u).value = dic_destino[numoperacion]
		sheet.cell(column=4, row=u).value = dic_bonificacion[numoperacion]
		sheet.cell(column=6, row=u).value = dic_factura[numoperacion]
		if dic_vendedor[numoperacion] == 1:
			sheet.cell(column=5, row=u).value = "Vallarta"
		else:
			sheet.cell(column=5, row=u).value = "Losada"
		labelll = Label(frame, text="Se ha guardado con exito")
		labelll.grid(column=3, row=3)
		prenum = Label(preframe, text= numoperacion)
		prenum.grid(column=1, row=u)
		predes = Label(preframe, text= dic_destino[numoperacion])
		predes.grid(column=2, row=u)
		prebon = Label(preframe, text= dic_bonificacion[numoperacion])
		prebon.grid(column=3, row=u)
		prefac = Label(preframe, text= dic_factura[numoperacion])
		prefac.grid(column=5, row=u)
		if dic_vendedor[numoperacion] ==1: 
			preven = Label(preframe, text= "Vallarta")
			preven.grid(column=4, row=u)
		else:
			preven = Label(preframe, text= "Losada")
			preven.grid(column=4, row=u)	
	pen = (len(sheet['A']) - 1)
	label_envio_num.config(text=pen)
	wb.save('flexcontrol.xlsx')
	numero_operacion.delete(0, 'end')
	numero_factura.delete(0, 'end')
def preview():
	sheet = wb[str(today)]
	ent = (len(sheet['A'])+1)
	for e in range(1, ent):
		prenum = Label(preframe, text= sheet.cell(column=1, row=e).value)
		prenum.grid(column=1, row=e)
		predes = Label(preframe, text= sheet.cell(column=3, row=e).value)
		predes.grid(column=2, row=e)
		prebon = Label(preframe, text= sheet.cell(column=4, row=e).value)
		prebon.grid(column=4, row=e)
		preven = Label(preframe, text= sheet.cell(column=5, row=e).value)
		preven.grid(column=4, row=e)
		prefac = Label(preframe, text= sheet.cell(column=6, row=e).value)
		prefac.grid(column=5, row=e)	
	pen = (len(sheet['A']) - 1)
	label_envio_num.config(text=pen)
	wb.save('flex.xlsx')


#GUI
raiz = Tk()
raiz.title("Libreria Losada")

frame = Frame(raiz)
frame.pack()

titulo = Label(frame, text="Libreria Losada \n Flex control")
titulo.grid(column=2, row=1)

mframe = Frame(frame)
mframe.grid(column=2, row=2)
vendframe = Frame(frame)
vendframe.grid(column=1, row=2)

opcion = IntVar()
ml = Radiobutton(mframe, text="Mercado Libre", variable=opcion, value=1)
ml.pack()
pr = Radiobutton(mframe, text="Propio", variable=opcion, value=2)
pr.pack()
opcion.set(1)

vendedor = IntVar()
vallarta = Radiobutton(vendframe, text="Vallarta", variable=vendedor, value=1)
vallarta.pack()
losada = Radiobutton(vendframe, text="Losada", variable=vendedor, value=2)
losada.pack()
vendedor.set(2)

label_noperacion = Label(frame, text="Número de operación:")
label_noperacion.grid(column=1, row=3)
numero_operacion = Entry(frame)
numero_operacion.grid(column=2, row=3)

label_factura = Label(frame, text= "Número de factura:")
label_factura.grid(column=1, row=4)
numero_factura = Entry(frame)
numero_factura.grid(column=2, row=4)

label_destino = Label(frame, text="Destino:")
label_destino.grid(column=1, row=5)
lugar = ttk.Combobox(frame, values=['CABA', 'GBA 1', 'GBA 2'])
lugar.grid(column=2, row=5)
lugar.current(0)

chk_bonificado_state = BooleanVar()
chk_bonificado = Checkbutton(frame, text='Bonificado:', var=chk_bonificado_state)
chk_bonificado.grid(column=2, row=6)

guardar = Button(frame, text="Guardar", command=botonguardar)
guardar.grid(column=2, row=7)

eframe = Frame(frame)
eframe.grid(column=3, row=2)

label_envio = Label(eframe, text="Envíos pendientes:")
label_envio.pack()
label_envio.config(font=('Helvetica 18 bold', 20))
label_envio_num = Label(eframe, text=pen)
label_envio_num.config(font=(30))
label_envio_num.pack()

botoncalcular = Button(frame, text="Calcular", command=ventanacalcular)
botoncalcular.grid(column=1, row=6)

botonpreview = Button(frame, text="Vista previa", command=preview)
botonpreview.grid(column=3, row=6)

preframe = Frame(raiz)
preframe.pack()


raiz.mainloop()
