#agarrador de precios

import openpyxl as op
import pandas as pd
import requests as req

libro_cambios=op.load_workbook("cambios.xlsx")
hoja=libro_cambios["Hoja2"]
eml = pd.read_csv(r"C:/Users/David/Desktop/EML.txt", sep=';',error_bad_lines=False)

largo=len(hoja['A'])
otrlargo=1

for linea in range(1,largo+1):
	if linea%2 > 0:
		isbn=str(hoja.cell(row=linea,column=1).value)
		print(isbn)
		try:
			fila= eml[eml.ISBN==isbn]
			
			print(otrlargo)
			precio=fila.iat[0,28]
			hoja.cell(row=otrlargo,column=4).value=fila.iat[0,1]
			print(fila.iat[0,1])
			hoja.cell(row=otrlargo,column=5).value=fila.iat[0,2]
			hoja.cell(row=otrlargo,column=6).value=fila.iat[0,3]
			hoja.cell(row=otrlargo,column=7).value=precio
			hoja.cell(row=otrlargo,column=3).value=isbn
			#print(precio)
			otrlargo+=1
		except IndexError:
			print("no en la base")
	else:
		nprecio=hoja.cell(row=linea,column=1).value
		hoja.cell(row=otrlargo-1,column=8).value=nprecio
libro_cambios.save("cambios.xlsx")
	