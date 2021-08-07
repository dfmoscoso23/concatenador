#Flexcontrol2.1
from tkinter import*
from tkinter import ttk
from tkinter import messagebox as mb
import openpyxl as op
from datetime import date, timedelta, datetime
from fpdf import FPDF
import requests
import obtenerTokenmodulo as otm

#valores:
flexcaba=250
flexgba1=400
flexgba2=500

mercadocaba=379.99
mercadogba1=599.99
mercadogba2=819.99

tok_losada=False
tok_vallarta=False
dic_envio={}
dic_traduccion={
	'cancelled':"cancelado",
	'shipped':"en camino",
	'delivered':"entregado",
	'ready_to_ship':"En preparación",
	'pending':"pendiente"
}
dic_posicionexcel={}
dic_cambiodestino={}
fr2=False
lis_fact=[]
lam_vendedor = lambda x: "Losada" if x ==str(553279780) else"Vallarta"
wb = op.load_workbook('flexcontrol.xlsx')

dic_vgenerales={'CABA':0,'GBA 1':0,'GBA 2':0}
dic_precios={'CABA':mercadocaba,'GBA 1':mercadogba1,'GBA 2':mercadogba2}

def pretoken():
	global tok_losada
	tok_losada=otm.revisar_token_activo("losada")
	if tok_losada==False:
		otm.obtenertoken("losada")
	else:	
		pass

	global tok_vallarta
	tok_vallarta=otm.revisar_token_activo("vallarta")
	if tok_vallarta==False:
		otm.obtenertoken("vallarta")
		#pretoken()
	else:	
		pass
def cargarventas(tok):
	vendedor=otm.obtener_vendedor(tok)
	fecha=date.today()
	ayer=fecha+timedelta(days=-1)
	if ayer.weekday() == 6:
		ayer=ayer+timedelta(days=-2)	
	fecha_a=ayer.strftime("%Y-%m-%d")
	fecha_b=fecha.strftime("%Y-%m-%d")
	url="https://api.mercadolibre.com/orders/search?seller="+str(vendedor)+"&order.date_created.from="+str(fecha_a)+"T12:00:00.000-04:00&order.date_created.to="+str(fecha_b)+"T12:00:00.000-04:00"
	headers={'Authorization': 'Bearer '+tok}
	r = requests.get(url, headers=headers)
	if r.status_code == 200:
		rjson=r.json()
		results=rjson['results']
		for item in results:
			ident_orden=item['id']
			shipping=item['shipping']
			url="https://api.mercadolibre.com/shipments/"+str(shipping['id'])
			headers={'Authorization': 'Bearer '+tok}
			r2 = requests.get(url, headers=headers)
			if r2.status_code == 200:
				rjson2=r2.json()
				logistic=rjson2['logistic_type']
				if logistic == 'self_service':
					option=rjson2['shipping_option']
					costo=option['cost']
					costo_base=option['list_cost']
					direccion=rjson2['receiver_address']
					provincia=direccion['state']['name']
					ciudad=direccion['city']['name']
					descuento=rjson2['cost_components']['loyal_discount']
					status=rjson2['status']
					if costo != costo_base:
						costo_final=option['list_cost']
						bonificado=True
					elif costo_base==0:
						costo_final=0
					else:	
						porcentaje=1-descuento
						costo_final=(costo_base)/porcentaje	
						bonificado=False
					dic_envio[shipping['id']]=[costo_final,bonificado,vendedor,ident_orden,provincia,ciudad,status]
					if status != 'cancelled':
						guardar_excel(shipping['id'],costo_final,bonificado,vendedor,ident_orden)		
def guardar_excel(nopera,destino,bonificado,vendedor,orden):
	fecha=date.today()
	hojas = wb.sheetnames
	if str(fecha) not in hojas:
		wb.create_sheet(str(fecha))
		sheet = wb[str(fecha)]
		sheet['A1'].value = 'Número de operación'
		sheet['B1'].value = 'Tipo'
		sheet['C1'].value = 'Destino'
		sheet['D1'].value = 'Bonificado'
		sheet['E1'].value = 'Vendedor'
		sheet['F1'].value = 'Factura'
		sheet['G1'].value = 'orden'
	else:
		sheet = wb[str(fecha)]
	guardado=False
	u = (sheet.max_row + 1)
	for x in range(1, u):
		xrow = sheet.cell(column=1, row=x).value
		if str(xrow) == str(nopera):
			if sheet.cell(column=3, row=x).value == None:
				dic_envio[nopera][0]=0
			elif sheet.cell(column=3, row=x).value == "CABA":
				dic_envio[nopera][0]=309.99
			elif sheet.cell(column=3, row=x).value == "GBA 1":
				dic_envio[nopera][0]=479.99
			else:
				dic_envio[nopera][0]=649.99
			dic_posicionexcel[nopera]=x		
			guardado=True
	if guardado==False:
		sheet.cell(column=1, row=u).value = nopera	
		sheet.cell(column=2, row=u).value = 1
		if destino==0:
			sheet.cell(column=3, row=u).value = None
		elif destino<310:
			sheet.cell(column=3, row=u).value = "CABA"
		elif destino <480:
			sheet.cell(column=3, row=u).value = "GBA 1"
		else:
			sheet.cell(column=3, row=u).value = "GBA 2"	
		sheet.cell(column=4, row=u).value = bonificado
		if vendedor == str(553279780):
			sheet.cell(column=5, row=u).value = "Losada"
		else:
			sheet.cell(column=5, row=u).value = "Vallarta"
		sheet.cell(column=7, row=u).value = orden
		dic_posicionexcel[nopera]=u	
	wb.save('flexcontrol.xlsx')
def contarenvios():
	fecha=date.today()
	sheet = wb[str(fecha)]
	u = (sheet.max_row - 1)
	label_contar_tit=Label(encabezado,text="Envíos del día:")
	label_contar_tit.grid(column=3,row=2)
	label_contar=Label(encabezado,text=u)
	label_contar.grid(column=4,row=2)
def deslplegarenvios():
	global frame2
	lis_fact.clear()
	fecha=date.today()
	sheet = wb[str(fecha)]
	frame2=Frame(frame_activo)
	frame2.pack(fill=BOTH, expand=1)
	canvas = Canvas(frame2)
	canvas.pack(side=LEFT, fill=BOTH, expand=1)
	scroll = ttk.Scrollbar(frame2, orient=VERTICAL, command=canvas.yview)
	scroll.pack(side=RIGHT, fill=Y)
	#scroll2 = ttk.Scrollbar(frame2, orient=HORIZONTAL, command=canvas.xview)
	#scroll2.pack(side=BOTTOM, fill=X)
	canvas.configure(yscrollcommand=scroll.set)
	canvas.bind('<Configure>', lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
	global frame_de_envios
	frame_de_envios = Frame(canvas)
	canvas.create_window((0,0), window=frame_de_envios, anchor="nw")
	z=0
	for item in dic_envio:
		desplegados=[]
		label_numenvio=Label(frame_de_envios,text=item)
		label_numenvio.grid(column=0,row=z)
		label_numorden=Label(frame_de_envios,text=dic_envio[item][3])
		label_numorden.grid(column=1,row=z)
		label_provincia=Label(frame_de_envios,text=dic_envio[item][4])
		label_provincia.grid(column=2,row=z)
		label_ciudad=Label(frame_de_envios,text=dic_envio[item][5])
		label_ciudad.grid(column=3,row=z)	
		if dic_envio[item][0] == 0:
			entry_costo= Entry(frame_de_envios, width=4)
			entry_costo.grid(column=4,row=z)
			dic_cambiodestino[z]=entry_costo   
		else:	
			label_costo=Label(frame_de_envios,text='${:,.2f}'.format(dic_envio[item][0]))
			label_costo.grid(column=4,row=z)
		label_bonificado=Label(frame_de_envios,text=dic_envio[item][1])
		label_bonificado.grid(column=5,row=z)
		entry_factura= Entry(frame_de_envios, width=7)
		entry_factura.grid(column=6,row=z)
		fact=sheet.cell(column=6, row=dic_posicionexcel[item]).value
		if fact != None:
			entry_factura.insert(END,fact)
		lis_fact.append(entry_factura)
		label_vendedor=Label(frame_de_envios,text=lam_vendedor(dic_envio[item][2]))
		label_vendedor.grid(column=7,row=z)
		label_status=Label(frame_de_envios,text=dic_traduccion[dic_envio[item][-1]])
		label_status.grid(column=8,row=z)
		z+=1
def botoncargarventas():
	cargarventas(tok_vallarta)
	cargarventas(tok_losada)
	contarenvios()
	global fr2
	if fr2==True:
		global frame2
		frame2.destroy()
	deslplegarenvios()
	fr2=True
def guardar():
	fecha=date.today()
	hojas = wb.sheetnames
	if str(fecha) not in hojas:
		mb.showerror("Error","Primero es necesario cargar ventas")
	else:
		sheet = wb[str(fecha)]	
	global frame_de_envios
	b=2
	for item in lis_fact:
		num_fact=item.get()
		sheet.cell(column=6, row=b).value = num_fact
		if num_fact != "":
			item.destroy()
			label_fact = Label(frame_de_envios, text=num_fact)
			label_fact.grid(column=6,row=(b-2))
		else:
			pass
		try:
			destino=dic_cambiodestino[b-2].get()
			if destino !="":
				if float(destino)==0:
					sheet.cell(column=3, row=b).value = None
				elif float(destino)<310:
					sheet.cell(column=3, row=b).value = "CABA"
				elif float(destino)<480:
					sheet.cell(column=3, row=b).value = "GBA 1"
				else:
					sheet.cell(column=3, row=b).value = "GBA 2"
				dic_cambiodestino[b-2].destroy()
				label_dest = Label(frame_de_envios, text='${:,.2f}'.format(float(destino)))
				label_dest.grid(column=4,row=(b-2))
				dic_envio[sheet.cell(column=1, row=b).value][0]=destino
				del dic_cambiodestino[b-2]	
		except KeyError:
			pass				
		b+=1
		wb.save('flexcontrol.xlsx')
	lis_fact.clear()	
	mb.showinfo("Aviso","Guardado Satisfactoriamente")	
def modificar():
	global lis_fact
	lis_fact.clear()
	fecha=date.today()
	hojas = wb.sheetnames
	if str(fecha) not in hojas:
		mb.showerror("Error","Primero es necesario cargar ventas")
	else:
		sheet = wb[str(fecha)]
	global frame_de_envios
	z=0
	for item in dic_envio:
		fact=sheet.cell(column=6, row=(z+2)).value
		entry_factura= Entry(frame_de_envios, width=7)
		entry_factura.grid(column=6,row=z)
		entry_factura.insert(END,fact)
		lis_fact.append(entry_factura)
		z+=1
def agregarenviopropio():
	global agregar_top
	agregar_top = Toplevel(raiz)
	agregar_frame=Frame(agregar_top)
	agregar_frame.pack()
	nfact_label =Label(agregar_frame, text="Número de factura:")
	nfact_label.grid(column=1, row=1)
	global nfact_entry
	nfact_entry = Entry(agregar_frame, width=8)
	nfact_entry.grid(column=2, row=1)
	destino_label =Label(agregar_frame, text="Destino:")
	destino_label.grid(column=1, row=2)
	global lugar
	lugar = ttk.Combobox(agregar_frame, values=['CABA', 'GBA 1', 'GBA 2'])
	lugar.grid(column=2, row=2)
	lugar.current(0)
	vendedor_label =Label(agregar_frame, text="Vendedor:")
	vendedor_label.grid(column=1, row=3)
	global vendedor_combo
	vendedor_combo = ttk.Combobox(agregar_frame, values=['Losada', 'Vallarta'])
	vendedor_combo.grid(column=2, row=3)
	vendedor_combo.current(0)
	boton_guardar_envpropio = Button(agregar_frame, text="Guardar", command=guardarenviopropio)
	boton_guardar_envpropio.grid(column=2,row=4)
def guardarenviopropio():
	global nfact_entry
	nfact=nfact_entry.get()
	global lugar
	luga = lugar.get()
	if luga == "CABA":
		valor=309.99
	elif luga == "GBA 1":
		valor=479.99
	else:
		valor=649.99
	global vendedor_combo
	combo = vendedor_combo.get()
	if combo == "Losada":
		vend=str(553279780)
	else:
		vend="Vallarta"	
	dic_envio[nfact]=[int(valor),False,vend,nfact,luga,luga,"ready_to_ship"]	
	guardar_excel(nfact,valor,False,vend,nfact)
	contarenvios()
	global frame2
	frame2.destroy()
	deslplegarenvios()
	global agregar_top
	agregar_top.destroy()
def ventanacalcular():
	today = date.today()
	global calc
	calc = Toplevel(raiz)
	aframe = Frame(calc)
	aframe.grid(column=0,row=0)
	global ingre_fecha
	ingre_fecha_label = Label(aframe, text="Último día de la semana: (AAAA-MM-DD):")
	ingre_fecha_label.grid(column=1, row=1)
	ingre_fecha = Entry(aframe, width=10)
	ingre_fecha.insert(END, today)
	ingre_fecha.grid(column=2, row=1)
	global check_vallarta_var
	check_vallarta_var = BooleanVar()
	check_vallarta=Checkbutton(aframe,text="Vallarta",var=check_vallarta_var)
	check_vallarta.grid(column=0, row=2)
	global check_losada_var
	check_losada_var = BooleanVar()
	check_losada=Checkbutton(aframe,text="Losada",var=check_losada_var)
	check_losada.grid(column=0, row=3)
	global rframe
	rframe = Frame(calc)
	rframe.grid(column=0,row=1)
	resul_label =Label(rframe, text="Resultados:")
	resul_label.grid(column=1, row=1)
	resul_caba_label =Label(rframe, text="Cantidad de envíos en CABA:")
	resul_caba_label.grid(column=1, row=2)
	resul_gba1_label =Label(rframe, text="Cantidad de envíos en GBA 1:")
	resul_gba1_label.grid(column=1, row=3)
	resul_gba2_label =Label(rframe, text="Cantidad de envíos en GBA 2:")
	resul_gba2_label.grid(column=1, row=4)
	resul_paga_label =Label(rframe, text="Cantidad a pagar:")
	resul_paga_label.grid(column=1, row=5)
	resul_paga2_label =Label(rframe, text="Cantidad de dinero recibido:")
	resul_paga2_label.grid(column=1, row=6)
	resul_paga3_label =Label(rframe, text="Saldo a pagar:")
	resul_paga3_label.grid(column=1, row=7)

	boton_calcular = Button(aframe, text="Calcular", command=precalcular)
	boton_calcular.grid(column=1, row=2)
	boton_detalles = Button(aframe, text="Detalles", command=precalcular2)
	boton_detalles.grid(column=2, row=2)
	boton_exportar = Button(aframe, text="Exportar Informe", command=exportarinforme)
	boton_exportar.grid(column=3, row=2)
def precalcular():
	global ingre_fecha
	fecharaw = ingre_fecha.get()
	year, month, day = map(int, fecharaw.split('-'))
	fecha = date(year, month, day)
	global check_losada_var
	global check_vallarta_var
	vendlosada = check_losada_var.get()
	vendvallarta = check_vallarta_var.get()
	if vendlosada == True and vendvallarta==False:
		vendedorc='Losada'
	elif vendlosada == False and vendvallarta==True:
		vendedorc='Vallarta'
	else:
		vendedorc='nada'
	calcular(fecha,vendedorc)
	dic_vgenerales['CABA']=0
	dic_vgenerales['GBA 1']=0
	dic_vgenerales['GBA 2']=0
def calculo(u, sheet, destino, vendedorc):
	lis_dest = []
	lis_bon=[]
	lis_ent=[]
	lis_prop=[]
	for x in range(1,u):
		if sheet.cell(column=5, row=x).value==vendedorc:
			xdestino=sheet.cell(column=3, row=x).value
			xbonificado = sheet.cell(column=4, row=x).value
			xtipo = sheet.cell(column=2, row=x).value
			if xdestino == destino:
				lis_dest.append('x')
				if xtipo == 2:
					lis_prop.append('x')
				else:
					if xbonificado==True:
						lis_bon.append('x')
					else:
						lis_ent.append('x')
	global drecibido
	drecibido += float((int(len(lis_bon)) * round(dic_precios[destino]/2))+(int(len(lis_ent))*dic_precios[destino]))
	dic_vgenerales[destino]+=len(lis_dest)
def detallador(u,sheet,destino,vendedorc):
	lis_dest=[]
	for x in range(1,u):
		if sheet.cell(column=5, row=x).value==vendedorc:
			xdestino=sheet.cell(column=3, row=x).value
			if xdestino == destino:
				lis_dest.append('x')
	return len(lis_dest)			
def calcular(fecha, vendedor):
	ini = fecha
	fin = ini + timedelta(days=-7)
	global drecibido
	drecibido = int(0)

	try:
		while ini >= fin:
			sheet = wb[str(ini)]
			u = (sheet.max_row + 1)
			calculo(u,sheet,'CABA',vendedor)
			calculo(u,sheet,'GBA 1',vendedor)
			calculo(u,sheet,'GBA 2',vendedor)
			ini = ini + timedelta(days=-1)
	except KeyError:
		pass
	pagaraflex = float((dic_vgenerales['CABA']*flexcaba)+(dic_vgenerales['GBA 1']*flexgba1)+(dic_vgenerales['GBA 2']*flexgba2))
	totalapagar = float(pagaraflex-drecibido)
	global rframe
	bell2 = Label(rframe, text=dic_vgenerales['CABA'])
	bell2.grid(column=2, row=2)
	bell = Label(rframe, text=dic_vgenerales['GBA 1'])
	bell.grid(column=2, row=3)
	bell3 = Label(rframe, text=dic_vgenerales['GBA 2'])
	bell3.grid(column=2, row=4)
	bell4 = Label(rframe, text='${:,.2f}'.format(pagaraflex))
	bell4.grid(column=2, row=5)
	bell5 = Label(rframe, text='${:,.2f}'.format(drecibido))
	bell5.grid(column=2, row=6)
	bell6 = Label(rframe, text= '${:,.2f}'.format(totalapagar))
	bell6.grid(column=2, row=7)
def precalcular2():
	global ingre_fecha
	fecharaw = ingre_fecha.get()
	year, month, day = map(int, fecharaw.split('-'))
	fecha = date(year, month, day)
	global check_losada_var
	global check_vallarta_var
	vendlosada = check_losada_var.get()
	vendvallarta = check_vallarta_var.get()
	if vendlosada == True and vendvallarta==False:
		vendedorc='Losada'
	elif vendlosada == False and vendvallarta==True:
		vendedorc='Vallarta'
	else:
		vendedorc='nada'
	detalles(fecha,vendedorc)
def detalles(fecha,vendedorc):
	global calc
	dframe = Frame(calc)
	dframe.grid(column=0, row=3)
	inis = []
	ini = fecha
	fin = ini + timedelta(days=-7)
	global drecibido
	drecibido = int(0)
	try:
		while ini >= fin:
			inis.append('x')
			l = len(inis)
			sheet = wb[str(ini)]
			u = (sheet.max_row + 1)
			detaf = Label(dframe, text=str(ini))
			detaf.grid(column=(l+1), row=1)
			deta = Label(dframe, text=detallador(u,sheet,'CABA',vendedorc))
			deta.grid(column=(l+1), row=2)
			deta2 = Label(dframe, text=detallador(u,sheet,'GBA 1',vendedorc))
			deta2.grid(column=(l+1), row=3)
			deta3 = Label(dframe, text=detallador(u,sheet,'GBA 2',vendedorc))
			deta3.grid(column=(l+1), row=4)
			ini = ini + timedelta(days=-1)
	except KeyError:
		pass
	deta4 = Label(dframe, text="CABA")
	deta4.grid(column=1, row=2)
	deta5 = Label(dframe, text="GBA1")
	deta5.grid(column=1, row=3)
	deta6 = Label(dframe, text="GBA2")
	deta6.grid(column=1, row=4)
def exportarinforme():
	inis=[]
	global ingre_fecha
	fecharaw = ingre_fecha.get()
	year, month, day = map(int, fecharaw.split('-'))
	fecha = date(year, month, day)
	ini = fecha
	fin = ini + timedelta(days=-7)
	global check_losada_var
	global check_vallarta_var
	vendlosada = check_losada_var.get()
	vendvallarta = check_vallarta_var.get()
	if vendlosada == True and vendvallarta==False:
		vendedorc='Losada'
	elif vendlosada == False and vendvallarta==True:
		vendedorc='Vallarta'
	else:
		vendedorc='nada'
	pdf = FPDF(orientation='L', format ='A4')
	pdf.add_page()
	pdf.set_font("Arial",'B', size=15)
	pdf.text((5),(5), txt="Librería Losada - Informe de envíos por Flex")
	global drecibido
	drecibido = int(0)
	rx=25
	try:
		while ini >= fin:
			pdf.set_font("Arial", size=10)
			sheet = wb[str(ini)]
			u = (sheet.max_row + 1)
			calculo(u,sheet,'CABA',vendedorc)
			calculo(u,sheet,'GBA 1',vendedorc)
			calculo(u,sheet,'GBA 2',vendedorc)
			inis.append('x')
			l = len(inis)
			sheet = wb[str(ini)]
			u = (sheet.max_row + 1)
			pdf.text(rx,80,txt=str(ini))
			pdf.text(rx+5,85,txt=str(detallador(u,sheet,'CABA',vendedorc)))
			pdf.text(rx+5,90,txt=str(detallador(u,sheet,'GBA 1',vendedorc)))
			pdf.text(rx+5,95,txt=str(detallador(u,sheet,'GBA 2',vendedorc)))		
			rx+=20
			ini = ini + timedelta(days=-1)
	except KeyError:
		pass
	pagaraflex = float((dic_vgenerales['CABA']*flexcaba)+(dic_vgenerales['GBA 1']*flexgba1)+(dic_vgenerales['GBA 2']*flexgba2))
	totalapagar = float(pagaraflex-drecibido)	
	pdf.set_font("Arial",'B', size=15)	
	pdf.text(5,15,txt="Vendedor:")
	pdf.text(45,15, txt=str(vendedorc))
	pdf.text(5,25,txt="Cantidad de envíos en CABA:")
	pdf.text(85,25, txt=str(dic_vgenerales['CABA']))
	pdf.text(5,30,txt="Cantidad de envíos en GBA 1:")
	pdf.text(85,30, txt=str(dic_vgenerales['GBA 1']))
	pdf.text(5,35,txt="Cantidad de envíos en GBA 2:")
	pdf.text(85,35, txt=str(dic_vgenerales['GBA 2']))
	pdf.text(5,40,txt="Monto a pagar a Flex:")
	pdf.text(65,40, txt=str('${:,.2f}'.format(pagaraflex)))
	pdf.text(5,45,txt="Dinero recibido:")
	pdf.text(65,45, txt=str('${:,.2f}'.format(drecibido)))
	pdf.text(5,50,txt="Total:")
	pdf.text(65,50, txt=str('${:,.2f}'.format(totalapagar)))
	pdf.text(5,75, txt=">>>>>>>>>>>>>>>>>DETALLE>>>>>>>>>>>>>>>>>>>>>")
	pdf.text(5,85,txt="CABA:")
	pdf.text(5,90,txt="GBA 1:")
	pdf.text(5,95,txt="GBA 2:")	
	ini=fecha
	pdf.add_page('L')
	pdf.text(5,5, txt=">>>>>>>>>>>>>>>>>OPERACIÓN ML vs FACTURA>>>>>>>>>>>>>>>>>>>>>")
	sx=5
	try:
		while ini >= fin:
			sheet = wb[str(ini)]
			pdf.set_font("Arial", size=7)
			pdf.text(sx,10,txt=str(ini))
			lar=15
			for ope in range(1,u):
				if ope == 1:
					pdf.text(sx,lar,txt=str(sheet.cell(column=1, row=ope).value)+">"+str(sheet.cell(column=6, row=ope).value)+">"+str(sheet.cell(column=3, row=ope).value))
					lar+=5
				elif sheet.cell(column=5, row=ope).value==vendedorc:
					pdf.text(sx,lar,txt=str(sheet.cell(column=1, row=ope).value)+">"+str(sheet.cell(column=6, row=ope).value)+">"+str(sheet.cell(column=3, row=ope).value))
					lar+=5		
			sx+=45
			rx+=20
			ini = ini + timedelta(days=-1)
	except KeyError:
		pass
	try:			
		pdf.output('informeenviosflexsemana'+vendedorc+str(fecha)+'.pdf','F')
		mb.showinfo("Aviso","Exportado con éxito")
	except PermissionError:
		mb.showerror("ERROR","PDF abierto")

	dic_vgenerales['CABA']=0
	dic_vgenerales['GBA 1']=0
	dic_vgenerales['GBA 2']=0						
#GUI
raiz = Tk()
raiz.title("Libreria Losada")
raiz.geometry('900x300')
encabezado = Frame(raiz)
encabezado.pack()
titulo = Label(encabezado, text="Libreria Losada \n Flex control")
titulo.grid(column=2, row=1)

botonera = Frame(raiz)
botonera.pack()

boton_otoken = Button(botonera, text="Obtener Token", command=pretoken)
boton_otoken.grid(column=0,row=0)
boton_cventas = Button(botonera, text="Cargar ventas", command=botoncargarventas)
boton_cventas.grid(column=1,row=0)
boton_guardarc = Button(botonera, text="Guardar Cambios", command=guardar)
boton_guardarc.grid(column=2,row=0)
boton_modificarc = Button(botonera, text="Modificar", command=modificar)
boton_modificarc.grid(column=3,row=0)
boton_calcular = Button(botonera, text="Agregar envío propio", command=agregarenviopropio)
boton_calcular.grid(column=4,row=0)
boton_calcular = Button(botonera, text="Calcular", command=ventanacalcular)
boton_calcular.grid(column=5,row=0)

frame_activo = Frame(raiz,width=750, height=300)
frame_activo.pack()
frame_activo.pack_propagate(False)

raiz.mainloop()