#Preba matcheador

import openpyxl as op
libro_publica="ISB"
hoja_publica="Hoja1"
columna_ml=1
columna_isbn_publi=2
lista_ML=[]
dic_isbn={}
dic_ml={}

def matchear_ml_isbn():
	global libro_publica
	libro1 = libro_publica
	global hoja_publica
	hoja1 = hoja_publica
	lista_publicaciones = op.load_workbook(libro1 +'.xlsx')
	hoja_publicaciones = lista_publicaciones[hoja1]
	ga = (len(hoja_publicaciones['A'])+1)
	for x in range (2,ga):
		global columna_ml
		item = hoja_publicaciones.cell(row=x, column=int(columna_ml)).value 
		global columna_isbn_publi
		isbn = hoja_publicaciones.cell(row=x, column=int(columna_isbn_publi)).value
		if isbn != None:
			lista_ML.append(item)
			dic_isbn[item]=isbn
			dic_ml[str(isbn)]=item
	print(len(dic_ml))


def printeable():
	isbnbuscado=input("Ingrese ISBN")
	print(dic_ml[isbnbuscado])
matchear_ml_isbn()
printeable()