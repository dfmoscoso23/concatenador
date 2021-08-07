#Transferidor
import requests
import json
import pandas as pd
import re
import openpyxl as op

eml = pd.read_csv(r"C:/Users/David/Desktop/EML.txt", sep=';',error_bad_lines=False)

def descargar_datos(mla):
	token="APP_USR-4726037063911819-080414-63d3ad66aa886ae5ef2f9b9a41dc31a2-553279780"
	url="https://api.mercadolibre.com/items/"+mla
	headers = {"Authorization": str("Bearer "+token),"Content-type": "application/json" }

	req = requests.get(url, headers=headers)
	print(req)
	if req.status_code == 200:
		regjson=req.json()
		imagenes=regjson['pictures']
		atributos=regjson['attributes']
		sku=atributos[-1]['value_name']
		imag=[]
		for item in imagenes:
			imag.append({'source':item['url']})
		return(sku,imag)	

def recolectar_datos(sku,mla):
	fila= eml[eml.ISBN==sku]
	pretitulo=fila.iat[0,1]
	preautor=fila.iat[0,2]
	editorial=(fila.iat[0,3]).strip()
	if re.search(r'VARIOS',editorial) !=None:
		editorial=" "
	elif re.search(r'ASIGNACION',editorial) !=None:
		editorial=" "
	preprecio=str(fila.iat[0,28])
	reprecio=preprecio.split('.')
	precio=reprecio[0]
	tema=(fila.iat[0,5]).strip()
	stock=fila.iat[0,17]
	proveedor=fila.iat[0,4]

	#imag2=["https://i.postimg.cc/bJL0FhCh/9789504002826001.jpg","https://i.postimg.cc/R07LKfwq/9789504002826002.jpg","https://i.postimg.cc/B6SMfSSh/001.jpg"]
	return (pretitulo, preautor,precio,tema, stock, proveedor, editorial)
def corrector_titulo(pretitulo):
	retitulo=re.search(r"(.*), ([LE][A-Z]\w?)(.*)", pretitulo)
	if retitulo == None:
		prepre=pretitulo
	else:
		titulo=(retitulo[2]+" "+retitulo[1]+" "+retitulo[3]).strip()
		prepre=titulo
	retitulo2=re.search(r"([\w 0-9]*)(/L)([\w 0-9]*)",prepre)
	if retitulo2 == None:
		return prepre.strip()
	else:
		titulo=(retitulo2[1]+retitulo2[3]).strip()
		return titulo
def corrector_autor(preautor):	
	reautor = re.search(r"(.*), (.*)",preautor)
	if reautor == None:
		return (preautor.strip(),preautor.strip())
	else:	
		autor=reautor[2]+" "+reautor[1]
		apellido=reautor[1]
		return (autor, apellido)	
def condicional_stock(stock,editorial):
	if stock<=0:
		return False
	elif re.search("ALIANZA",editorial) != None:
			return False
	elif re.search("CATEDRA",editorial) != None:
			return False
	elif re.search("TECNOS",editorial) != None:
		return False
	elif re.search("FACTORIA",editorial) != None:
		return False
	elif re.search("CARALT",editorial) != None:
		return False				
	return True
def condicional_losada(editorial):
	if re.search("LOSADA",editorial) != None:
		return False
	if re.search("AIQUE",editorial) != None:
		return False
	if re.search("LAROUSSE",editorial) != None:
		return False	
	return True	
def pausar(mla):
	token3="APP_USR-4726037063911819-080317-d637c2eb94c31e52d75565cc25ac8443-553279780"
	url3 = "https://api.mercadolibre.com/items/" + mla
	headers3 = {'Authorization' : ('Bearer '+ token3), 'Content-type': 'application/json', 'Accept':'application/json'}
	payload = {"status": 'paused'}
	response3 = requests.put(url3, headers=headers3, json=payload)
	print(response3.status_code)
def control_repeticion(sku):
	hoja_lista=excel_lista["VAL"]
	cantidad=len(hoja_lista["A"])
	for x in range(1,cantidad+1):
		xrow=hoja_lista.cell(row=x,column=2).value
		if xrow == int(sku):
			return False
	return True		
def eje(mla):
	sku,imag=descargar_datos(mla)
	try:
		pretitulo, preautor,precio,tema, stock, proveedor, editorial=recolectar_datos(sku, mla)	
		titulo=corrector_titulo(pretitulo)
		autor,apellido=corrector_autor(preautor)
		titulo_publicacion=(titulo+" - "+apellido+" - "+editorial).strip()
		if condicional_stock(stock,editorial)==True:
			titulo=corrector_titulo(pretitulo)
			autor,apellido=corrector_autor(preautor)
			titulo_publicacion=(titulo+" - "+apellido+" - "+editorial).strip()
			if control_repeticion(sku):
				newid=subir_mercadolibre(titulo_publicacion,precio,imag,autor,tema,titulo,sku,editorial)
				control_cambiados= open("cambiados_3.txt",'a')
				control_cambiados.write(str(sku)+";"+str(mla)+";"+str(newid)+"\n")#+";"+str(titulo_publicacion)+";"+str(titulo)+";"+str(autor)+";"+str(editorial)+";"+str(stock)+";"+str(precio)
				control_cambiados.close()
				if condicional_losada(editorial) == True:
					#pausar(mla)
					control_pausados= open("pausados_3.txt",'a')
					control_pausados.write(str(sku)+";"+str(mla)+"\n")
					control_pausados.close()
				else:
					control_losadas= open("losadas_3.txt",'a')
					control_losadas.write(str(sku)+";"+str(mla)+"\n")
					control_losadas.close()
		else:
			if stock>0:
				control_alianzas= open("alianzas_3.txt",'a')
				control_alianzas.write(str(sku)+";"+str(mla)+"\n")
				control_alianzas.close()
			else:
				control_sinstock= open("sinstock_3.txt",'a')
				control_sinstock.write(str(sku)+";"+str(mla)+"\n")
				control_sinstock.close()	
	except IndexError:
		control_errores= open("errores_3.txt",'a')
		control_errores.write(str(sku)+";"+str(mla)+"\n")
		control_errores.close()	
def subir_mercadolibre(titulo_publicacion,precio,imag,autor,tema,titulo,sku,editorial):			
	data= {
	  "title":str(titulo_publicacion[:59]),
	  "category_id":"MLA412445",
	  "price":precio,
	  "currency_id":"ARS",
	  "available_quantity":1,
	  "buying_mode":"buy_it_now",
	  "condition":"new",
	  "listing_type_id":"gold_special",
	  "sale_terms":[
	     {
	        "id":"WARRANTY_TYPE",
	        "value_name":"Garantía del vendedor"
	     },
	     {
	        "id":"WARRANTY_TIME",
	        "value_name":"30 días"
	     }
	  ],
	  "pictures":imag,
	  "shipping":{
			  	"mode":"me2",
			  	"local_pick_up": True,
			  	"logistic_type": "xd_drop_off"
		},
	  "attributes":[
	     {
	        "id":"AUTHOR",
	        "value_name":autor
	     },
	     {
	        "id":"BOOK_GENRE",
	        "value_name":tema
	     },
	     {
	        "id":"BOOK_TITLE",
	        "value_name":titulo
	     },
	     {
	        "id":"FORMAT",
	        "value_name":"Papel"
	     },
	     #{
	       # "id":"GTIN",
	        #"value_name":sku
	     #},
	     {
	        "id":"ITEM_CONDITION",
	        "value_name":"Nuevo"
	     },
	     {
	        "id":"LANGUAGE",
	        "value_name":"Español"
	     },
	     {
	        "id":"NARRATION_TYPE",
	        "value_name":"Auto ayuda"
	     },
	     {
	        "id":"PUBLISHER",
	        "value_name":editorial
	     },
	     {
	        "id":"SELLER_SKU",
	        "value_name":sku
	     }
	  ]
	}
	token2="APP_USR-4726037063911819-080414-efd28f88bedf9f693e59447b3dc89e16-736684164"
	url2="https://api.mercadolibre.com/items/"
	headers2 = {"Authorization": str("Bearer "+token2)}
	req2 = requests.post(url2, headers=headers2, json=data)
	print(req2)
	if req2.status_code == 201:
		resp=req2.json()
		newid=resp['id']
		return newid
	else:
		print(req2.text)
		#raise ValueError("SERVIDOR CARGADO")	
#print(recolectar_datos("MLA885273386"))
#'MLA901546104''MLA863059279','MLA911191665'

lista_prueba=[
"MLA923631495",
"MLA923630626",
"MLA923630625",
"MLA860427877",
"MLA909256448",
"MLA895890645",
"MLA876569194",
"MLA884037094"
]
excel_lista=op.load_workbook("TITUcambio.xlsx")
hoja_lista=excel_lista["Hoja7"]
cantidad=len(hoja_lista["A"])
print(cantidad)
for x in range(1,cantidad+1):
	item=hoja_lista.cell(row=x,column=2).value
	eje(item)