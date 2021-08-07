#Eliminador
import re
import openpyxl as op
import requests as req
import pandas as pd

eml = pd.read_csv(r"C:/Users/David/Desktop/EML.txt", sep=';',error_bad_lines=False)

def preparador():
	excel=op.load_workbook("eliminables.xlsx")
	hoja=excel['Hoja1']
	cantidad=len(hoja['A'])
	for x in range(1,cantidad+1):
		try:
			sku=str(hoja.cell(row=x, column=1).value)
		except TypeError:
			break
		fila= eml[eml.ISBN==sku]
		try:
			pretitulo=fila.iat[0,1]
			preautor=fila.iat[0,2]
			editorial=(fila.iat[0,3]).strip()
			hoja.cell(row=x, column=4).value=pretitulo
			hoja.cell(row=x, column=5).value=preautor
			hoja.cell(row=x, column=6).value=editorial
		except IndexError:
			hoja.cell(row=x, column=4).value="no"
			hoja.cell(row=x, column=5).value="no"
			hoja.cell(row=x, column=6).value="no"
	excel.save("eliminables.xlsx")	