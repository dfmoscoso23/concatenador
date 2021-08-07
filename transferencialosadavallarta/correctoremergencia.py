#Corrector de emergencia
import re
import requests
import pandas as pd
import openpyxl as op

eml = pd.read_csv(r"C:/Users/David/Desktop/EML.txt", sep=';',error_bad_lines=False)

def recolectar_datos(sku):
	fila= eml[eml.ISBN==sku]
	pretitulo=fila.iat[0,1]
	preautor=fila.iat[0,2]
	editorial=(fila.iat[0,3]).strip()
	if re.search(r'VARIOS',editorial) !=None:
		editorial=" "
	elif re.search(r'ASIGNACION',editorial) !=None:
		editorial=" "
	preprecio=str(fila.iat[0,28])
	reprecio=preprecio.split('.')
	precio=reprecio[0]
	tema=(fila.iat[0,5]).strip()
	stock=fila.iat[0,17]
	proveedor=fila.iat[0,4]

	#imag2=["https://i.postimg.cc/bJL0FhCh/9789504002826001.jpg","https://i.postimg.cc/R07LKfwq/9789504002826002.jpg","https://i.postimg.cc/B6SMfSSh/001.jpg"]
	return (pretitulo, preautor,precio,tema, stock, proveedor, editorial)
def corrector_titulo(pretitulo):
	retitulo=re.search(r"(.*), ([LE][A-Z]\w?)(.*)", pretitulo)
	if retitulo == None:
		prepre=pretitulo
	else:
		titulo=(retitulo[2]+" "+retitulo[1]+" "+retitulo[3]).strip()
		prepre=titulo
	retitulo2=re.search(r"([\w ]*)(/L)([\w 0-9]*)",prepre)
	if retitulo2 == None:
		return prepre.strip()
	else:
		titulo=(retitulo2[1]+retitulo2[3]).strip()
		return titulo
def corrector_autor(preautor):	
	reautor = re.search(r"(.*), (.*)",preautor)
	if reautor == None:
		return (preautor.strip(),preautor.strip())
	else:	
		autor=reautor[2]+" "+reautor[1]
		apellido=reautor[1]
		return (autor, apellido)
def postear(mla,titulo):
	token3="APP_USR-4726037063911819-080318-625dc572aa6ded7e482cb5b2923674a4-736684164"
	url3 = "https://api.mercadolibre.com/items/" + mla
	headers3 = {'Authorization' : ('Bearer '+ token3), 'Content-type': 'application/json', 'Accept':'application/json'}
	payload = {"title": titulo[:59]}
	response3 = requests.put(url3, headers=headers3, json=payload)
	print(response3.status_code)		
def eje(sku,mla):
	pretitulo, preautor,precio,tema, stock, proveedor, editorial=recolectar_datos(str(sku))	
	titulo=corrector_titulo(pretitulo)
	autor,apellido=corrector_autor(preautor)
	titulo_publicacion=(titulo+" - "+apellido+" - "+editorial).strip()
	print(titulo_publicacion[:59])
	postear(mla,titulo_publicacion)

excel_lista=op.load_workbook("TITUcambio.xlsx")
hoja_lista=excel_lista["Hoja4"]
cantidad=len(hoja_lista["A"])
print(cantidad)
for x in range(1,cantidad+1):
	sku=hoja_lista.cell(row=x,column=1).value
	mla=hoja_lista.cell(row=x,column=3).value
	eje(sku,mla)