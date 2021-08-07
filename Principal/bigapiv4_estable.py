#BIGAPI V4
import openpyxl as op
import requests
import json
from datetime import date
from tkinter import*
from tkinter import ttk
from tkinter import messagebox
from datetime import date

#TOKEN
today = date.today()
lista_items = []
tok=""
usuario = 553279780
revisor_desdeExcel = False
revisor_desdeAPI= False
#Cargar catálogo
libro_publica=""
hoja_publica=""
columna_ml=""
columna_isbn_publi=""
#Precios
dic_preciosml = {}
dic_isbn = {}
dic_preciosultra = {}
lista_ML = []
preciosbajos = []
dic_titulo = {}
dic_autor = {}
list_var = []
dic_preciosbajos = {}
lis_precio = []
lis_fuera = []
lis_cambiados = []
lis_cambiados2 = []
lis_erroresdevacio = []
lis_erroresdeapi = []
lis_procesados = []
#stock
dic_stockml = {}
dic_statusml = {}
dic_isbn_stock = {}
dic_stockultra = {}
lista_ML_stock = []
stockbajos = []
stockaltos = []
stockaltospausados = []
dic_titulo_stock = {}
dic_autor_stock = {}
dic_var_stock = {}
list_var_stock = []
list_var_pausa = []
dic_var_pausa = {}
dic_stockbajos = {}
dic_editorial_stock = {}
dic_cambio_stock = {}
dic_cambios={}
lis_stock = []
lis_fuera_stock = []
lis_cambiados_stock = []
lis_cambiados2_stock = []
lis_erroresdevacio_stock = []
lis_erroresdeapi_stock = []
lis_procesados_stock = []
#titulo
dic_nuevotitulo={}
dic_editorial={}
dic_isbn_titulo = {}
lista_ML_titulo = []
dic_titulo_titulo = {}
dic_autor_titulo = {}
lis_fuera_titulo = []
lis_cambiados_titulo = []
lis_erroresdevacio_titulo = []
lis_erroresdeapi_titulo = []
lis_procesados_titulo = []
global sa
sa=0
global sl
sl=50
global d
d=0
global h
h=1

#Obtener Token
def solicitar(url, data, headers):
	response = requests.post(url, headers=headers, json=data)
	if response.status_code == 200:
		resp = str(response.status_code)
		respcont = str(response.content)
		resptext = str(response.text)
		respjson = response.json()
		global tok
		tok=respjson['access_token']
		st_label_rtoken.config(text="Cargado")
		token_box.insert(END, tok)
		f = open('token.txt', "a")
		f.write(resp)
		f.write("\t")
		f.write(respcont)
		f.write("\n")
		f.write(resptext)
		f.close()
	else:
		messagebox.showerror(message="No se pudo obtener el Token, revise el código TG")	
def botonear():
	url = "https://api.mercadolibre.com/oauth/token"
	headers ={'accept': 'application/json','content-type': 'application/x-www-form-urlencoded'}
	data = {
	'grant_type':'authorization_code',
	'client_id':'4726037063911819',
	'client_secret':'SKjt3ZUGtiXM90wvOn2xlvraWXEvQH2N',
	'redirect_uri':'https://localhost:30000'
	}	
	global tgbox
	cod = tgbox.get()
	data['code']=cod
	solicitar(url, data, headers)
def obtenertoken():
	#Toplevel
	toplevel_obtener_token = Toplevel(raiz)
	toplevel_obtener_token.title("Obtener Token")
	frame_OT = Frame(toplevel_obtener_token)
	frame_OT.pack()

	copiar_label=Label(frame_OT, text="Copie esto en su navegador para obtener la autorización:")
	copiar_label.grid(column=2, row=1, padx=5, pady=5)
	copiar_entry = Entry(frame_OT, width=75)
	copiar_entry.grid(column=2, row=2, padx=5, pady=5)
	copiar_entry.insert(END, "https://auth.mercadolibre.com.ar/authorization?response_type=code&client_id=4726037063911819&state=TRE7412&redirect_uri=https://localhost:30000")
	
	label1 = Label(frame_OT, text="Ingrese TG code:")
	label1.grid(column=1, row=3, padx=5, pady=5)
	global tgbox
	tgbox = Entry(frame_OT, width=75)
	tgbox.grid(column=2, row=3, padx=5, pady=5)

	boton = Button(frame_OT, text="Obtener Token", command=botonear)
	boton.grid(column=2, row=4, padx=5, pady=5)

	label1 = Label(frame_OT, text="Token:")
	label1.grid(column=1, row=4, padx=5, pady=5)
	global token_box
	token_box = Text(frame_OT, height=5, width=65)
	token_box.grid(column=2, row=5, padx=5, pady=5)
#Cargar catálogo
def catalogoenexcel():
	global libro_publica
	libro_publica = excel_publicaciones_libro.get()
	global hoja_publica
	hoja_publica= excel_publicaciones_hoja.get()
	global columna_ml
	columna_ml=excel_publicaciones_columna_ml_ent.get()
	global columna_isbn_publi
	columna_isbn_publi=excel_publicaciones_columna_isbn_ent.get()
	global revisor_desdeExcel
	revisor_desdeExcel=True
	messagebox.showinfo(message="Catálogo guardado")
	global toplevel_desdeExcel
	toplevel_desdeExcel.destroy()
	global st_label_rcatalogo	
	st_label_rcatalogo.config(text="Cargado desde Excel")
def desdeExcel():
	#Toplevel de ingreso de datos
	global toplevel_desdeExcel
	toplevel_desdeExcel = Toplevel(raiz)
	toplevel_desdeExcel.title("Ingresar Catálogo desde Excel")
	frame_DE = Frame(toplevel_desdeExcel)
	frame_DE.pack()
	frame_DE2=Frame(toplevel_desdeExcel)
	frame_DE2.pack()
	#Ingresar lista de publicaciones
	excel_publicaciones_libro_lab = Label(frame_DE, text="Inserte Libro de publicaciones:")
	excel_publicaciones_libro_lab.grid(column=1, row=2)
	excel_publicaciones_libro_lab2 = Label(frame_DE, text=".xlsx")
	excel_publicaciones_libro_lab2.grid(column=3, row=2)
	global excel_publicaciones_libro
	excel_publicaciones_libro = Entry(frame_DE, width=15)
	excel_publicaciones_libro.grid(column=2, row=2, padx=5, pady=5)
	excel_publicaciones_hoja_lab = Label(frame_DE, text="Inserte nombre de la hoja:")
	excel_publicaciones_hoja_lab.grid(column=1, row=3)
	global excel_publicaciones_hoja
	excel_publicaciones_hoja = Entry(frame_DE, width=15)
	excel_publicaciones_hoja.grid(column=2, row=3, padx=5, pady=5)
	excel_publicaciones_columna_ml_lab= Label(frame_DE, text="Inserte el número de columna de código ML:")
	excel_publicaciones_columna_ml_lab.grid(column=1, row=4)
	global excel_publicaciones_columna_ml_ent
	excel_publicaciones_columna_ml_ent= Entry(frame_DE, width=5)
	excel_publicaciones_columna_ml_ent.grid(column=2, row=4, padx=5, pady=5)
	excel_publicaciones_columna_isbn_lab= Label(frame_DE, text="Inserte el número de columna de ISBN:")
	excel_publicaciones_columna_isbn_lab.grid(column=1, row=5)
	global excel_publicaciones_columna_isbn_ent
	excel_publicaciones_columna_isbn_ent= Entry(frame_DE, width=5)
	excel_publicaciones_columna_isbn_ent.grid(column=2, row=5, padx=5, pady=5)
	bot_guardar_catalogo = Button(frame_DE2, text="Guardar", command=catalogoenexcel)
	bot_guardar_catalogo.pack()
def escribiendo():
	cat =open("cat"+str(today)+".txt", 'w')
	for it in lista_items:
		cat.write(it['id']+"; "+str(it['isbn'])+"; "+str(it['precio'])+"; "+str(it['cantidad'])+"; "+str(it['titulo']))
		cat.write("\n")
	cat.close()	
	messagebox.showinfo(message="Exportado satisfactoriamente")	
def catalogueando(token,resultado, headers):
	try:
		id_resultado = resultado['id']
		urli = "https://api.mercadolibre.com/items/"+id_resultado
		response2 = requests.get(urli, headers=headers)
		response2_json = response2.json()
		atributos = response2_json['attributes']
		at=(len(atributos)-1)
		atributos_sku = atributos[at]
		strisbn = atributos_sku['value_name']
		try:
			isbn=int(strisbn)
		except ValueError:
			isbn=1234567891011	
		precio = response2_json['price']
		cantidad = response2_json['available_quantity']
		titulo = response2_json['title']
		dic ={
		'id':id_resultado,
		'isbn':isbn,
		'precio':precio,
		'cantidad':cantidad,
		'titulo':titulo
		}
		lista_items.append(dic)
	except KeyError:
		pass		
def iterando(token, lresult, result, headers):
	for item in range(lresult):
		resultado = result[item]
		catalogueando(token, resultado, headers)	
def desdeAPI():
	if tok == "":
		messagebox.showerror(message="Es necesario cargar un Token previamente")
	else:	
		#Toplevel
		toplevel_desdeAPI = Toplevel(raiz)
		toplevel_desdeAPI.title("Cargar Catálogo desde API")
		frame_DA = Frame(toplevel_desdeAPI)
		frame_DA.pack()
		DA_label=Label(frame_DA, text="Procesando, esto puede tardar un momento")
		DA_label.pack()
		barra_api = ttk.Progressbar(frame_DA, orient=HORIZONTAL, mode='determinate')
		barra_api.pack()
		bot_exportar = Button(frame_DA, text="Exportar en CSV", command=escribiendo)
		bot_exportar.pack()
		url= "https://api.mercadolibre.com/sites/MLA/search?seller_id="+str(usuario)
		token = tok
		headers = {'Authorization': ('Bearer '+ token), 'Content-type': 'application/json', 'Accept':'application/json'}
		response = requests.get(url, headers=headers)
		if response.status_code == 200:
			response_json = json.loads(response.text)
			prepaginas = response_json['paging']
			paginas = round((prepaginas['total']/prepaginas['limit'])+0.5)
			valor_barra = (100/int(paginas))
			#result = response_json['results']
			#lresult = len(result)
			#iterando(token, lresult, result, headers)
			for pag in range(int(paginas)):
				pagurl= "https://api.mercadolibre.com/sites/MLA/search?seller_id="+str(usuario)+"&limit=50&offset="+str(pag*50) #&page=+str(pag)
				pagresponse = requests.get(pagurl, headers=headers)
				pagresponse_json = json.loads(pagresponse.text)
				result = pagresponse_json['results']
				lresult = len(result)
				iterando(token, lresult, result, headers)
				barra_api['value'] = valor_barra
				toplevel_desdeAPI.update_idletasks()
		messagebox.showinfo(message="Proceso concluido")
		global st_label_rcatalogo	
		st_label_rcatalogo.config(text="Cargado desde API")
		global revisor_desdeAPI
		revisor_desdeAPI=True

def cerrarcargarcatalogo():
	global frame_CC
	frame_CC.destroy()
	bot_cargar_catalogo.config(command=cargarcatalogo)		
def cargarcatalogo():
	#Toplevel
	#toplevel_cargar_catalogo = Toplevel(raiz)
	#toplevel_cargar_catalogo.title("Cargar Catálogo")
	global frame_CC
	frame_CC = Frame(frame_activo)
	frame_CC.pack()
	bot_desde_excel = Button(frame_CC, text="Desde una planilla de Excel", command=desdeExcel)
	bot_desde_excel.grid(column=1, row=1, padx=5, pady=5)
	bot_desde_api = Button(frame_CC, text="Desde Api", command=desdeAPI)
	bot_desde_api.grid(column=2, row=1, padx=5, pady=5)
	bot_cargar_catalogo.config(command=cerrarcargarcatalogo)
#Actualizar precios
def comparar(mlitem, token, frame3, barra_actualizar):
	try:	
		u = int(dic_preciosultra[mlitem])
		m = int(dic_preciosml[mlitem])
		barra_actualizar['value'] = 1
		#global toplevel_actualizador
		frame_activo.update_idletasks()
		global inlabelpro
		inlabelpro.config(text=str(len(lis_procesados)))
		if u > m:
			url2 = "https://api.mercadolibre.com/items/" + str(mlitem)
			headers2 = {'Authorization' : ('Bearer '+ token), 'Content-type': 'application/json', 'Accept':'application/json'}
			payload = {"price": u}
			response2 = requests.put(url2, headers=headers2, json=payload)
			if response2.status_code != 200:
				#messagebox.showerror(message=str(mlitem) + "No se pudo actualizar \n Error: " + str(response2.status_code) , title="Error de Api" )
				lis_erroresdeapi.append(mlitem)
			else:
				lis_cambiados.append(mlitem)
		if u < m:
			preciosbajos.append('X')
			d = int(len(preciosbajos))
			isb = dic_isbn[mlitem]
			titu = dic_titulo[mlitem]
			#auto = dic_autor[mlitem]
			mlitem2 = Label(frame3, text=mlitem)
			mlitem2.grid(column=1, row=d)
			nisbn= Label(frame3, text=isb)
			nisbn.grid(column=2, row=d)
			ntitulo =Label(frame3, text=titu)
			ntitulo.grid(column=3, row=d)
			#nautor =Label(frame3, text=auto)
			#nautor.grid(column=4, row=d)
			uprecio=Label(frame3, text=u)
			uprecio.grid(column=5, row=d)
			mprecio = Entry(frame3, width=5)
			mprecio.grid(column=6, row=d)
			mprecio.insert(END, str(m))
			lis_precio.append(mprecio)
			
			var = IntVar()
			Checkbutton(frame3, text="actualizar", variable=var).grid(row=d, column=7)
			list_var.append(var)
			de =d-1
			dic_preciosbajos[de]= mlitem

	except KeyError:
		lis_fuera.append(mlitem)
def isbeniador(ml, token, frame3, hoja_precios, columna_isbn, columa_precio, barra_actualizar):
	try:
		fa = (len(hoja_precios['A'])+1)
		for y in range(2,fa):
			isbnm = dic_isbn[ml]
			isbnu = hoja_precios.cell(row=y, column=int(columna_isbn)).value
			if str(isbnm) == str(isbnu):
				dic_preciosultra[ml]=hoja_precios.cell(row=y, column=int(columa_precio)).value
				#tit = hoja_precios.cell(row=y, column=2).value
				#dic_titulo[ml]= str(tit)
				#aut = hoja_precios.cell(row=y, column=3).value
				#dic_autor[ml]= str(aut)
		comparar(ml, token, frame3, barra_actualizar)
	except KeyError:
		lis_erroresdevacio.append(ml)
def diccionador(libro2, hoja2, columna_isbn, columa_precio, barra_actualizar):
	lista_precios = op.load_workbook(libro2 +'.xlsx')
	hoja_precios = lista_precios[hoja2]
	#desplegable de precios bajos
	baj = Toplevel(raiz)
	baj.geometry('950x500')
	frame4=Frame(baj)
	frame4.pack()
	label2 = Label(frame4, text="Libros con precio más bajo en la base de datos. ¿Desea actualizarlos?")
	label2.grid(column=1, row=1)
	check_but = Button(frame4, text="Actualizar", command=check)
	check_but.grid(column=2, row=1, padx=5, pady=5)	
	
	frame2=Frame(baj)
	frame2.pack(fill=BOTH, expand=1)

	canvas = Canvas(frame2)
	canvas.pack(side=LEFT, fill=BOTH, expand=1)

	scroll = ttk.Scrollbar(frame2, orient=VERTICAL, command=canvas.yview)
	scroll.pack(side=RIGHT, fill=Y)

	canvas.configure(yscrollcommand=scroll.set)
	canvas.bind('<Configure>', lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

	frame3 = Frame(canvas)
	canvas.create_window((0,0), window=frame3, anchor="nw")
	#token
	token = tok
	global revisor_desdeExcel
	global revisor_desdeAPI
	if revisor_desdeExcel == True:
		global libro_publica
		libro1 = libro_publica
		global hoja_publica
		hoja1 = hoja_publica
		lista_publicaciones = op.load_workbook(libro1 +'.xlsx')
		hoja_publicaciones = lista_publicaciones[hoja1]
		ga = (len(hoja_publicaciones['A'])+1)
		global info
		inlabelga = Label(info, text=" / "+str(ga))
		inlabelga.grid(column=3, row=2)
		for x in range (2,ga):
			global columna_ml
			item = hoja_publicaciones.cell(row=x, column=int(columna_ml)).value 
			global columna_isbn_publi
			isbn = hoja_publicaciones.cell(row=x, column=int(columna_isbn_publi)).value
			if isbn != None:
				lista_ML.append(item)
				dic_isbn[item]=isbn
				url = "https://api.mercadolibre.com/items/" + str(item)
				headers = {'Authorization': ('Bearer '+ token), 'Content-type': 'application/json', 'Accept':'application/json'}
				response = requests.get(url, headers=headers)
				if response.status_code == 200:
					jsonp = response.json()
					precio = jsonp['price']
					titul=jsonp['title']
					dic_titulo[item]=titul
					dic_preciosml[item]=precio
			lis_procesados.append(item)		
	# Precios del ultra
			isbeniador(item, token, frame3, hoja_precios, columna_isbn, columa_precio, barra_actualizar)		
	elif revisor_desdeAPI == True:
		ga = (len(lista_items))
		inlabelga = Label(info, text=" / "+str(ga))
		inlabelga.grid(column=3, row=1)
		for i in range(ga):
			dic = lista_items[i]
			item = dic['id']
			isbn = dic['isbn']
			dic_isbn[item]=isbn
			precio= dic['precio']
			dic_preciosml[item]=precio
			titulo = dic['titulo']
			dic_titulo[item]= str(titulo)
			isbeniador(item, token, frame3, hoja_precios, columna_isbn, columa_precio, barra_actualizar)

	else:
		messagebox.showerror(message="No está cargado el catálogo")				
	
	if len(lis_fuera) > 0:
		messagebox.showerror(message=str(len(lis_fuera)) + " registros, no se encuentran en la base de datos.", title="Error de Api" )
	if len(lis_erroresdevacio) > 0:
		messagebox.showerror(message=str(len(lis_erroresdevacio)) + "No se pudo actualizar por estar vacíos", title="Error" )
	if len(list_var) < 1:
		baj.destroy()	
	messagebox.showinfo(message=str(len(lis_cambiados))+ "fueron cambiados de precio", title="Acualización de precios")
	#except InvalidFileException:
		#messagebox.showerror(message="No hay una hoja de excel con precios")
def botonactualizar():
	excel_precios_libro
	libro2 = excel_precios_libro.get()
	hoja2 = excel_precios_hoja.get()
	columna_isbn=excel_precios_columna_isbn_ent.get()
	columa_precio=excel_precios_columna_precios_ent.get()
	barra_actualizar = ttk.Progressbar(info, orient=HORIZONTAL, mode="determinate")
	barra_actualizar.grid(column=1, row=1)
	diccionador(libro2, hoja2, columna_isbn, columa_precio, barra_actualizar)	
			#if tok == "":
		#messagebox.showerror(message="Es necesario cargar un Token previamente")
	#else:
	#Lista de Precios
def check():
	token = tok
	l=len(list_var)	
	for n in range(l):
		op = list_var[n].get()
		if op == 1:
			url3 = "https://api.mercadolibre.com/items/" + str(dic_preciosbajos[n])
			headers3 = {'Authorization' : ('Bearer '+ token), 'Content-type': 'application/json', 'Accept':'application/json'}
			p = int(lis_precio[n].get())
			payload = {"price": p}
			response2 = requests.put(url3, headers=headers3, json=payload)
			if response2.status_code != 200:
				messagebox.showerror(message=str(dic_preciosbajos[n]) + "No se pudo actualizar \n Error: " + str(response2.status_code) , title="Error de Api" )
				lis_erroresdeapi.append(dic_preciosbajos[n])
			else:
				lis_cambiados2.append(dic_preciosbajos[n])	
	messagebox.showinfo(message=str(len(lis_cambiados2))+ "fueron cambiados de precio", title="Acualización de precios")
def desplegadordevacio():
	desp1 = Toplevel(raiz)
	box = Text(desp1)
	box.pack()
	l = len(list_var)
	for n in range(l):
		box.insert(END, str(dic_isbn[dic_preciosbajos[n]]) + "\n" + str(dic_preciosml[dic_preciosbajos[n]])+ "\n")
def desplegadordeAPI():
	desp2 = Toplevel(raiz)
	box = Text(desp2)
	box.pack()
	box.insert(END, lis_erroresdeapi)
	box.insert(END, lis_erroresdevacio)
def desplegadorcambiados():
	desp3 = Toplevel(raiz)
	frame7=Frame(desp3)
	frame7.pack(fill=BOTH, expand=1)
	canvas = Canvas(frame7)
	canvas.pack(side=LEFT, fill=BOTH, expand=1)
	scroll = ttk.Scrollbar(frame7, orient=VERTICAL, command=canvas.yview)
	scroll.pack(side=RIGHT, fill=Y)
	canvas.configure(yscrollcommand=scroll.set)
	canvas.bind('<Configure>', lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
	frame8=Frame(canvas)
	canvas.create_window((0,0), window=frame8, anchor="nw")
	arbol = ttk.Treeview(frame8)
	arbol['columns']=("titulo", "isbn","idml","cambio")
	arbol.column("#0",width=5)
	arbol.column("titulo", anchor=W, width=240)
	arbol.column("isbn", anchor=CENTER, width=95)
	arbol.column("idml", anchor=W, width=95)
	arbol.column("cambio",anchor=W, width=80)
	arbol.heading("#0", text="Seleccionar", anchor=W)
	arbol.heading("titulo", text="Título", anchor=W)
	arbol.heading("isbn", text="ISBN", anchor=W)
	arbol.heading("idml", text="ID ML", anchor=W)
	arbol.heading("cambio", text="Cambio", anchor=W)
	arbol.grid(column=0,row=0)
	i=0
	for n in lis_cambiados2_stock:
		arbol.insert(parent="",index="end",iid=i,text="",value=(dic_titulo_stock[n],dic_isbn_stock[n],n,dic_cambios[n]))
		i+=1
	desp3 = Toplevel(raiz)
	box = Text(desp3)
	box.pack()
	if len(lis_cambiados) > 0:
		for m in lis_cambiados:
			box.insert(END, str(m)+" "+ str(dic_isbn[m])+" "+str(dic_titulo[m])+"\n")
	for n in lis_cambiados2:
		box.insert(END, str(n)+" "+ str(dic_isbn[n])+" "+str(dic_titulo[n])+"\n")
def cerraractualizadordeprecios():
	bot_actualizar_precios.config(command=actualizador_de_precios)
	global frame_AP
	frame_AP.destroy()

def actualizador_de_precios():
	#global toplevel_actualizador
	#toplevel_actualizador = Toplevel(raiz)
	#toplevel_actualizador.title("Actualizar Precios")
	bot_actualizar_precios.config(command=cerraractualizadordeprecios)
	global frame_AP
	frame_AP = Frame(frame_activo)
	frame_AP.pack()
	frame5= Frame(frame_AP)
	frame5.pack()
	frame6= Frame(frame_AP)
	frame6.pack()

	#Ingresar lista de precios
	excel_precios_libro_lab = Label(frame5, text="Inserte Libro de precios:")
	excel_precios_libro_lab.grid(column=1, row=2)
	excel_precios_libro_lab2 = Label(frame5, text=".xlsx")
	excel_precios_libro_lab2.grid(column=3, row=2)
	global excel_precios_libro
	excel_precios_libro = Entry(frame5, width=15)
	excel_precios_libro.grid(column=2, row=2, padx=5, pady=5)
	excel_precios_hoja_lab = Label(frame5, text="Inserte nombre de la hoja:")
	excel_precios_hoja_lab.grid(column=1, row=3)
	global excel_precios_hoja
	excel_precios_hoja = Entry(frame5, width=15)
	excel_precios_hoja.grid(column=2, row=3, padx=5, pady=5)
	excel_precios_columna_isbn_lab= Label(frame5, text="Inserte el número de columna de ISBN:")
	excel_precios_columna_isbn_lab.grid(column=1, row=4)
	global excel_precios_columna_isbn_ent
	excel_precios_columna_isbn_ent= Entry(frame5, width=5)
	excel_precios_columna_isbn_ent.grid(column=2, row=4, padx=5, pady=5)
	excel_precios_columna_precios_lab= Label(frame5, text="Inserte el número de columna de precio:")
	excel_precios_columna_precios_lab.grid(column=1, row=5)
	global excel_precios_columna_precios_ent
	excel_precios_columna_precios_ent= Entry(frame5, width=5)
	excel_precios_columna_precios_ent.grid(column=2, row=5, padx=5, pady=5)

	actualizar_precios = Button(frame6, text="Actualizar", command=botonactualizar)
	actualizar_precios.grid(column=2, row=4, padx=5, pady=5)
	global info
	info = Frame(frame_AP)
	info.pack()
	inlabel = Label(info, text="Registros que serán procesados: ")
	inlabel.grid(column=1, row=2)
	global inlabelpro
	inlabelpro = Label(info, text="0")
	inlabelpro.grid(column=2, row=2)
	botonera = Frame(frame_AP)
	botonera.pack()

	boton_de_error_de_vacio = Button(botonera, text="Precios bajos en Ultra", command=desplegadordevacio)
	boton_de_error_de_vacio.grid(column=1, row=3, padx=5, pady=5)
	boton_de_error_de_API = Button(botonera, text="Errores de API", command=desplegadordeAPI)
	boton_de_error_de_API.grid(column=2, row=3, padx=5, pady=5)
	boton_de_cambiados = Button(botonera, text="Precios Cambiados", command=desplegadorcambiados)
	boton_de_cambiados.grid(column=3, row=3, padx=5, pady=5)
#Controlar stocks

def desplegadordevacio_stock():
	desp1 = Toplevel(raiz)
	box = Text(desp1)
	box.pack()
	for n in stockbajos:
		box.insert(END, str(dic_isbn_stock[n]) + "\n" + str(dic_stockml[n])+ "\n")	
def desplegadordeAPI_stock():
	desp2 = Toplevel(raiz)
	box = Text(desp2)
	box.pack()
	box.insert(END, lis_erroresdeapi_stock)
	box.insert(END, lis_erroresdevacio_stock)
def desplegadorcambiados_stock():
	desp3 = Toplevel(raiz)
	frame7=Frame(desp3)
	frame7.pack(fill=BOTH, expand=1)
	canvas = Canvas(frame7)
	canvas.pack(side=LEFT, fill=BOTH, expand=1)
	scroll = ttk.Scrollbar(frame7, orient=VERTICAL, command=canvas.yview)
	scroll.pack(side=RIGHT, fill=Y)
	canvas.configure(yscrollcommand=scroll.set)
	canvas.bind('<Configure>', lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
	frame8=Frame(canvas)
	canvas.create_window((0,0), window=frame8, anchor="nw")
	arbol = ttk.Treeview(frame8)
	arbol['columns']=("titulo", "isbn","idml","cambio")
	arbol.column("#0",width=5)
	arbol.column("titulo", anchor=W, width=240)
	arbol.column("isbn", anchor=CENTER, width=95)
	arbol.column("idml", anchor=W, width=95)
	arbol.column("cambio",anchor=W, width=80)
	arbol.heading("#0", text="Seleccionar", anchor=W)
	arbol.heading("titulo", text="Título", anchor=W)
	arbol.heading("isbn", text="ISBN", anchor=W)
	arbol.heading("idml", text="ID ML", anchor=W)
	arbol.heading("cambio", text="Cambio", anchor=W)
	arbol.grid(column=0,row=0)
	i=0
	for n in lis_cambiados2_stock:
		arbol.insert(parent="",index="end",iid=i,text="",value=(dic_titulo_stock[n],dic_isbn_stock[n],n,dic_cambios[n]))
		i+=1
def check_stock():
	token = tok
	l=len(dic_var_stock)	
	for key in dic_var_stock:
		op = dic_var_stock[key].get()
		if op == 1:
			url3 = "https://api.mercadolibre.com/items/" + str(dic_stockbajos[key])
			headers3 = {'Authorization' : ('Bearer '+ token), 'Content-type': 'application/json', 'Accept':'application/json'}
			p = int(dic_cambio_stock[key].get())
			payload = {"available_quantity": p}
			response2 = requests.put(url3, headers=headers3, json=payload)
			if response2.status_code != 200:
				messagebox.showerror(message=str(dic_stockbajos[key]) + "No se pudo actualizar \n Error: " + str(response2.status_code) , title="Error de Api" )
				lis_erroresdeapi_stock.append(dic_stockbajos[key])
			else:
				lis_cambiados2_stock.append(dic_stockbajos[key])
				dic_cambios[key]="cantidad"	
	messagebox.showinfo(message=str(len(lis_cambiados2_stock))+ "fueron cambiados de stock", title="Acualización de stock")
def check_activar():
	token = tok	
	for key in dic_var_pausa:
		op = dic_var_pausa[key].get()
		if op == 1:
			url3 = "https://api.mercadolibre.com/items/" + str(dic_stockbajos[key])
			headers3 = {'Authorization' : ('Bearer '+ token), 'Content-type': 'application/json', 'Accept':'application/json'}
			payload = {"status": 'active'}
			response2 = requests.put(url3, headers=headers3, json=payload)
			if response2.status_code != 200:
				messagebox.showerror(message=str(dic_stockbajos[key]) + "No se pudo actualizar \n Error: " + str(response2.status_code) , title="Error de Api" )
				lis_erroresdeapi_stock.append(stockbajos[n])
			else:
				lis_cambiados2_stock.append(dic_stockbajos[key])
				dic_cambios[key]="activado"		
	messagebox.showinfo(message=str(len(lis_cambiados2_stock))+ "fueron cambiados de stock", title="Acualización de stock")
def check_pausar():
	token = tok	
	for key in dic_var_pausa:
		op = dic_var_pausa[key].get()
		if op == 1:
			url3 = "https://api.mercadolibre.com/items/" + str(dic_stockbajos[key])
			headers3 = {'Authorization' : ('Bearer '+ token), 'Content-type': 'application/json', 'Accept':'application/json'}
			payload = {"status": 'paused'}
			response2 = requests.put(url3, headers=headers3, json=payload)
			if response2.status_code != 200:
				messagebox.showerror(message=str(dic_stockbajos[key]) + "No se pudo actualizar \n Error: " + str(response2.status_code) , title="Error de Api" )
				lis_erroresdeapi_stock.append(dic_stockbajos[key])
			else:
				lis_cambiados2_stock.append(dic_stockbajos[key])
				dic_cambios[key]="pausado"		
	messagebox.showinfo(message=str(len(lis_cambiados2_stock))+ "fueron cambiados de stock", title="Acualización de stock")
def escribiendo_stockbajos():
	cat =open("stocksacambiarbajos"+str(today)+".txt", 'w')
	for it in stockbajos:
		cat.write(str(it)+"; "+str(dic_isbn_stock[it])+"; "+str(dic_titulo_stock[it])+"; "+str(dic_stockultra[it])+"; "+str(dic_stockml[it]))
		cat.write("\n")
	cat.close()	
	messagebox.showinfo(message="Exportado satisfactoriamente")
def escribiendo_stockaltos():
	cat =open("stocksacambiaraltos"+str(today)+".txt", 'w')
	for it in stockaltos:
		cat.write(str(it)+"; "+str(dic_isbn_stock[it])+"; "+str(dic_titulo_stock[it])+"; "+str(dic_stockultra[it])+"; "+str(dic_stockml[it]))
		cat.write("\n")
	cat.close()	
	messagebox.showinfo(message="Exportado satisfactoriamente")
def escribiendo_stockaltospausados():
	cat =open("stocksacambiaraltospausados"+str(today)+".txt", 'w')
	for it in stockaltospausados:
		cat.write(str(it)+"; "+str(dic_isbn_stock[it])+"; "+str(dic_titulo_stock[it])+"; "+str(dic_stockultra[it])+"; "+str(dic_stockml[it]))
		cat.write("\n")
	cat.close()	
	messagebox.showinfo(message="Exportado satisfactoriamente")	
def atrasbajos():
	dic_cambio_stock.clear()
	dic_var_stock.clear()
	dic_var_pausa.clear()
	dic_stockbajos.clear()
	global frame3
	frame3.destroy()
	global canvas
	frame3 = Frame(canvas)
	canvas.create_window((0,0), window=frame3, anchor="nw")
	global sa
	sa=sa-50
	global sl
	sl=sa-50
	for mlitem in stockbajos[sa:sl]:
		presentador(mlitem)
	global h
	h-1
	global hojas3
	hojas3.config(text=h)
def siguientebajos():
	dic_cambio_stock.clear()
	dic_var_stock.clear()
	dic_var_pausa.clear()
	dic_stockbajos.clear()
	global frame3
	frame3.destroy()
	global canvas
	frame3 = Frame(canvas)
	canvas.create_window((0,0), window=frame3, anchor="nw")
	global sa
	sa=sa+50
	global sl
	sl=sa+50
	for mlitem in stockbajos[sa:sl]:
		presentador(mlitem)
	global h
	h+1
	global hojas3
	hojas3.config(text=h)
def desplegarstockbajos():
	dic_cambio_stock.clear()
	dic_var_stock.clear()
	dic_var_pausa.clear()
	dic_stockbajos.clear()
	desp4 = Toplevel(raiz)
	desp4.geometry('1000x550')
	frame4=Frame(desp4)
	frame4.pack()
	label2 = Label(frame4, text="Libros con stock más bajo en la base de datos. ¿Desea actualizarlos?")
	label2.grid(column=1, row=1)
	check_but = Button(frame4, text="Actualizar stocks", command=check_stock)
	check_but.grid(column=2, row=1, padx=5, pady=5)
	checkpau_but = Button(frame4, text="Pausar", command=check_pausar)
	checkpau_but.grid(column=3, row=1, padx=5, pady=5)
	frame2=Frame(desp4)
	frame2.pack(fill=BOTH, expand=1)
	global canvas
	canvas = Canvas(frame2)
	canvas.pack(side=LEFT, fill=BOTH, expand=1)

	scroll = ttk.Scrollbar(frame2, orient=VERTICAL, command=canvas.yview)
	scroll.pack(side=RIGHT, fill=Y)

	canvas.configure(yscrollcommand=scroll.set)
	canvas.bind('<Configure>', lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
	global frame3
	frame3 = Frame(canvas)
	canvas.create_window((0,0), window=frame3, anchor="nw")
	exportar_but = Button(frame4, text="Exportar a CSV", command=escribiendo_stockbajos)
	exportar_but.grid(column=9, row=1, padx=5, pady=5)
	s=int(len(stockbajos))
	if s<50:
		for mlitem in stockbajos:
			presentador(mlitem)
	else:
		hojas=Label(frame4, text="Hoja:")
		hojas.grid(column=6, row=1)
		global hojas3
		hojas3=Label(frame4, text="1/")
		hojas3.grid(column=7, row=1)
		hojas2=Label(frame4, text=int((s/50)+1))
		hojas2.grid(column=8, row=1)
		atras_but = Button(frame4, text="Anteriores 50", command=atrasbajos)
		atras_but.grid(column=4, row=1, padx=5, pady=5)
		siguiente_but = Button(frame4, text="Siguiente 50", command=siguientebajos)
		siguiente_but.grid(column=5, row=1, padx=5, pady=5)
		for mlitem in stockbajos[sa:sl]:
			presentador(mlitem)
def atraspau():
	dic_cambio_stock.clear()
	dic_var_stock.clear()
	dic_var_pausa.clear()
	dic_stockbajos.clear()
	global frame3
	frame3.destroy()
	global canvas
	frame3 = Frame(canvas)
	canvas.create_window((0,0), window=frame3, anchor="nw")
	global sa
	sa=sa-50
	global sl
	sl=sa-50
	for mlitem in stockaltospausados[sa:sl]:
		presentador(mlitem)
	global h
	h-1
	global hojas3
	hojas3.config(text=h)
def siguientepau():
	dic_cambio_stock.clear()
	dic_var_stock.clear()
	dic_var_pausa.clear()
	dic_stockbajos.clear()
	global frame3
	frame3.destroy()
	global canvas
	frame3 = Frame(canvas)
	canvas.create_window((0,0), window=frame3, anchor="nw")
	global sa
	sa=sa+50
	global sl
	sl=sa+50
	for mlitem in stockaltospausados[sa:sl]:
		presentador(mlitem)
	global h
	h+1
	global hojas3
	hojas3.config(text=h)
def atras():
	dic_cambio_stock.clear()
	dic_var_stock.clear()
	dic_stockbajos.clear()
	global frame3
	frame3.destroy()
	global canvas
	frame3 = Frame(canvas)
	canvas.create_window((0,0), window=frame3, anchor="nw")
	global sa
	sa=sa-50
	global sl
	sl=sa-50
	for mlitem in stockaltos[sa:sl]:
		presentador(mlitem)
	global h
	h-1
	global hojas3
	hojas3.config(text=h)
def siguiente():
	dic_cambio_stock.clear()
	dic_var_stock.clear()
	dic_stockbajos.clear()
	global frame3
	frame3.destroy()
	global canvas
	frame3 = Frame(canvas)
	canvas.create_window((0,0), window=frame3, anchor="nw")
	global sa
	sa=sa+50
	global sl
	sl=sa+50
	for mlitem in stockaltos[sa:sl]:
		presentador(mlitem)
	global h
	h+1
	global hojas3
	hojas3.config(text=h)	
def presentadorbajos(mlitem):
	global d
	global frame3
	isb = dic_isbn_stock[mlitem]
	titu = dic_titulo_stock[mlitem]
	#auto = dic_autor[mlitem]
	#edit = dic_editorial[mlitem]
	mlitem2 = Label(frame3, text=mlitem)
	mlitem2.grid(column=1, row=d)
	nisbn= Label(frame3, text=isb)
	nisbn.grid(column=2, row=d)
	ntitulo =Label(frame3, text=titu)
	ntitulo.grid(column=3, row=d)
	#nautor =Label(frame3, text=auto)
	#nautor.grid(column=4, row=d)
	#nedit =Label(frame3, text=edit)
	#nedit.grid(column=5, row=d)
	ustock=Label(frame3,  text=str(dic_stockultra[mlitem]))
	ustock.grid(column=6, row=d)
	mstock = Entry(frame3, width=5)
	mstock.grid(column=7, row=d)
	mstock.insert(END, str(dic_stockml[mlitem]) )
	mstatu=Label(frame3, text=dic_statusml[mlitem])
	mstatu.grid(column=8, row=d)
	dic_cambio_stock[mlitem]=ustock
	var_stock = IntVar()
	vs = Checkbutton(frame3, text="stock", variable=var_stock)
	vs.grid(row=d, column=9)
	dic_var_stock[mlitem]=var_stock
	var_pausa = IntVar()
	Checkbutton(frame3, text="activar", variable=var_pausa).grid(row=d, column=10)
	dic_var_pausa[mlitem]=var_pausa
	de =d
	dic_stockbajos[mlitem]= mlitem
	d=d+1
def presentador(mlitem):
	global d
	global frame3
	isb = dic_isbn_stock[mlitem]
	titu = dic_titulo_stock[mlitem]
	mlitem2 = Label(frame3, text=mlitem)
	mlitem2.grid(column=1, row=d)
	nisbn= Label(frame3, text=isb)
	nisbn.grid(column=2, row=d)
	ntitulo =Label(frame3, text=titu)
	ntitulo.grid(column=3, row=d)
	mstock = Label(frame3, text=dic_stockml[mlitem])
	mstock.grid(column=6, row=d)
	mstatu=Label(frame3, text=dic_statusml[mlitem])
	mstatu.grid(column=7, row=d)
	ustock=Entry(frame3, width=5)
	ustock.grid(column=8, row=d)
	ustock.insert(END, str(dic_stockultra[mlitem]))
	dic_cambio_stock[mlitem]=ustock
	var_stock = IntVar()
	vs = Checkbutton(frame3, text="stock", variable=var_stock)
	vs.grid(row=d, column=9)
	dic_var_stock[mlitem]=var_stock
	var_pausa = IntVar()
	Checkbutton(frame3, text="pausar", variable=var_pausa).grid(row=d, column=10)
	dic_var_pausa[mlitem]=var_pausa
	de =d
	dic_stockbajos[mlitem]= mlitem
	d=d+1
def desplegastocksaltospausados():
	dic_cambio_stock.clear()
	dic_var_stock.clear()
	dic_var_pausa.clear()
	dic_stockbajos.clear()
	desp5 = Toplevel(raiz)
	desp5.geometry('1000x550')
	frame4=Frame(desp5)
	frame4.pack()
	label2 = Label(frame4, text="Libros con stock más bajo en la base de datos. ¿Desea actualizarlos?")
	label2.grid(column=1, row=1)
	check_but = Button(frame4, text="Actualizar stocks", command=check_stock)
	check_but.grid(column=2, row=1, padx=5, pady=5)
	checkpau_but = Button(frame4, text="Activar", command=check_activar)
	checkpau_but.grid(column=3, row=1, padx=5, pady=5)
	frame2=Frame(desp5)
	frame2.pack(fill=BOTH, expand=1)
	global canvas
	canvas = Canvas(frame2)
	canvas.pack(side=LEFT, fill=BOTH, expand=1)

	scroll = ttk.Scrollbar(frame2, orient=VERTICAL, command=canvas.yview)
	scroll.pack(side=RIGHT, fill=Y)

	canvas.configure(yscrollcommand=scroll.set)
	canvas.bind('<Configure>', lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
	global frame3
	frame3 = Frame(canvas)
	canvas.create_window((0,0), window=frame3, anchor="nw")
	exportar_but = Button(frame4, text="Exportar a CSV", command=escribiendo_stockaltospausados)
	exportar_but.grid(column=9, row=1, padx=5, pady=5)
	s=int(len(stockaltospausados))
	if s<50:
		for mlitem in stockaltospausados:
			presentadorbajos(mlitem)
	else:
		hojas=Label(frame4, text="Hoja:")
		hojas.grid(column=6, row=1)
		global hojas3
		hojas3=Label(frame4, text="1/")
		hojas3.grid(column=7, row=1)
		hojas2=Label(frame4, text=int((s/50)+1))
		hojas2.grid(column=8, row=1)
		atras_but = Button(frame4, text="Anteriores 50", command=atraspau)
		atras_but.grid(column=4, row=1, padx=5, pady=5)
		siguiente_but = Button(frame4, text="Siguiente 50", command=siguientepau)
		siguiente_but.grid(column=5, row=1, padx=5, pady=5)
		for mlitem in stockaltospausados[sa:sl]:
			presentadorbajos(mlitem)	
def desplegarstockaltos():
	dic_cambio_stock.clear()
	dic_var_stock.clear()
	dic_var_pausa.clear()
	dic_stockbajos.clear()
	desp5 = Toplevel(raiz)
	desp5.geometry('1000x550')
	frame4=Frame(desp5)
	frame4.pack()
	label2 = Label(frame4, text="Libros con stock más bajo en la base de datos. ¿Desea actualizarlos?")
	label2.grid(column=1, row=1)
	check_but = Button(frame4, text="Actualizar stocks", command=check_stock)
	check_but.grid(column=2, row=1, padx=5, pady=5)
	checkpau_but = Button(frame4, text="Pausar", command=check_pausar)
	checkpau_but.grid(column=3, row=1, padx=5, pady=5)
	frame2=Frame(desp5)
	frame2.pack(fill=BOTH, expand=1)
	global canvas
	canvas = Canvas(frame2)
	canvas.pack(side=LEFT, fill=BOTH, expand=1)

	scroll = ttk.Scrollbar(frame2, orient=VERTICAL, command=canvas.yview)
	scroll.pack(side=RIGHT, fill=Y)

	canvas.configure(yscrollcommand=scroll.set)
	canvas.bind('<Configure>', lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
	global frame3
	frame3 = Frame(canvas)
	canvas.create_window((0,0), window=frame3, anchor="nw")
	exportar_but = Button(frame4, text="Exportar a CSV", command=escribiendo_stockaltos)
	exportar_but.grid(column=9, row=1, padx=5, pady=5)
	s=int(len(stockaltos))
	if s<50:
		for mlitem in stockaltos:
			presentador(mlitem)
	else:
		hojas=Label(frame4, text="Hoja:")
		hojas.grid(column=6, row=1)
		global hojas3
		hojas3=Label(frame4, text="1/")
		hojas3.grid(column=7, row=1)
		hojas2=Label(frame4, text=int((s/50)+1))
		hojas2.grid(column=8, row=1)
		atras_but = Button(frame4, text="Anteriores 50", command=atras)
		atras_but.grid(column=4, row=1, padx=5, pady=5)
		siguiente_but = Button(frame4, text="Siguiente 50", command=siguiente)
		siguiente_but.grid(column=5, row=1, padx=5, pady=5)
		for mlitem in stockaltos[sa:sl]:
			presentador(mlitem)
			
def comparar_stock(mlitem, token):
	try:	
		u = int(dic_stockultra[mlitem])
		if u == 0:
			url = "https://api.mercadolibre.com/items/" + str(mlitem)
			headers = {'Authorization': ('Bearer '+ token), 'Content-type': 'application/json', 'Accept':'application/json'}
			response = requests.get(url, headers=headers)
			if response.status_code == 200:
				jsonp = response.json()
				stock = jsonp['available_quantity']
				titul = jsonp['title']
				dic_titulo_stock[mlitem]=titul
				dic_stockml[mlitem]=stock
				m = int(dic_stockml[mlitem])
				stat = jsonp['status']
				dic_statusml[mlitem]=stat
				q = str(dic_statusml[mlitem])
				if q == 'active':
					stockbajos.append(mlitem)
					print("bajos")
					print(len(stockbajos))
		if u != 0:
			url = "https://api.mercadolibre.com/items/" + str(mlitem)
			headers = {'Authorization': ('Bearer '+ token), 'Content-type': 'application/json', 'Accept':'application/json'}
			response = requests.get(url, headers=headers)
			if response.status_code == 200:
				jsonp = response.json()
				stock = jsonp['available_quantity']
				titul = jsonp['title']
				dic_titulo_stock[mlitem]=titul
				dic_stockml[mlitem]=stock
				m = int(dic_stockml[mlitem])
				stat = jsonp['status']
				dic_statusml[mlitem]=stat
				q = str(dic_statusml[mlitem])
				if q != 'active':
					if int(stock) > 0:
						stockaltospausados.append(mlitem)
						print("pausados")
						print(len(stockaltospausados))
					else:	
						stockaltos.append(mlitem)
						print("altos")
						print(len(stockaltos))		
	except KeyError:
		lis_fuera_stock.append(mlitem)
def isbeniador_stock(ml, token, hoja_stock, columna_isbn_stock, columna_stock):
	try:
		fa = (len(hoja_stock['A'])+1)
		for y in range(2,fa):
			isbnm = dic_isbn_stock[ml]
			isbnu = hoja_stock.cell(row=y, column=int(columna_isbn_stock)).value
			if str(isbnm) == str(isbnu):
				dic_stockultra[ml]=hoja_stock.cell(row=y, column=int(columna_stock)).value
				#tit = hoja_stock.cell(row=y, column=2).value
				#dic_titulo_stock[ml]= str(tit)
				#aut = hoja_stock.cell(row=y, column=3).value
				#dic_autor[ml]= str(aut)
				#edi = hoja_stock.cell(row=y, column=4).value
				#dic_editorial[ml] = str(edi)
		comparar_stock(ml, token)
	except KeyError:
		lis_erroresdevacio_stock.append(ml)
def revisador(token, hoja_stock, columna_isbn_stock, columna_stock):
	if revisor_desdeExcel == True:
		global libro_publica
		libro1 = libro_publica
		global hoja_publica
		hoja1 = hoja_publica
		lista_publicaciones = op.load_workbook(libro1 +'.xlsx')
		hoja_publicaciones = lista_publicaciones[hoja1]
		ga = (len(hoja_publicaciones['A'])+1)
		inlabelga = Label(info2, text=" / "+str(ga))
		inlabelga.grid(column=3, row=2)
		for x in range (2,ga):
			global columna_ml
			item = hoja_publicaciones.cell(row=x, column=int(columna_ml)).value 
			global columna_isbn_publi
			isbn = hoja_publicaciones.cell(row=x, column=int(columna_isbn_publi)).value
			if isbn != None:	
				dic_isbn_stock[item]=isbn
				lista_ML_stock.append(item)
				isbeniador_stock(item, token, hoja_stock, columna_isbn_stock, columna_stock)
	elif revisor_desdeAPI == True:
		ga = (len(lista_items))
		inlabelga = Label(info2, text=" / "+str(ga))
		inlabelga.grid(column=3, row=1)
		for i in range(ga):
			dic = lista_items[i]
			item = dic['id']
			isbn = dic['isbn']
			dic_isbn_stock[item]=isbn
			cant_stock = dic['cantidad']
			dic_stockml[item]=cant_stock
			titulo = dic['titulo']
			dic_titulo_stock[item]= str(titulo)
			isbeniador_stock(item, token, hoja_stock, columna_isbn_stock, columna_stock)
	else:
		messagebox.showerror(message="No está cargado el catálogo")
def iniciador():
	#Lista de stock
	libro_stock = excel_stocks_libro.get()
	hoja_stocks = excel_stocks_hoja.get()
	columna_isbn_stock=excel_stocks_columna_isbn_ent.get()
	columna_stock=excel_stocks_columna_stocks_ent.get()
	lista_stock = op.load_workbook(libro_stock +'.xlsx')
	hoja_stock = lista_stock[hoja_stocks]
	
	#token
	token = tok

	revisador(token, hoja_stock, columna_isbn_stock,columna_stock)
			
	# stock del ultra
				
	if len(lis_fuera_stock) > 0:
		messagebox.showerror(message=str(len(lis_fuera_stock)) + " registros, no se encuentran en la base de datos.", title="Error de Api" )
	if len(lis_erroresdevacio_stock) > 0:
		messagebox.showerror(message=str(len(lis_erroresdevacio_stock)) + "No se pudo actualizar por estar vacíos", title="Error" )
	messagebox.showinfo(message="Proceso Concluido")
def cerrarcontrolstocks():
	bot_controlar_stock.config(command=controlstocks)
	global frame_CSS
	frame_CSS.destroy()
def controlstocks():
	bot_controlar_stock.config(command=cerrarcontrolstocks)
	global frame_CSS
	frame_CSS = Frame(frame_activo)
	frame_CSS.pack()
	frame_CS= Frame(frame_CSS)
	frame_CS.pack()
	frame_CS2= Frame(frame_CSS)
	frame_CS2.pack()

	#Ingresar lista de stocks
	excel_stocks_libro_lab = Label(frame_CS, text="Inserte Libro de stocks:")
	excel_stocks_libro_lab.grid(column=1, row=2)
	excel_stocks_libro_lab2 = Label(frame_CS, text=".xlsx")
	excel_stocks_libro_lab2.grid(column=3, row=2)
	global excel_stocks_libro
	excel_stocks_libro = Entry(frame_CS, width=15)
	excel_stocks_libro.grid(column=2, row=2, padx=5, pady=5) 
	excel_stocks_hoja_lab = Label(frame_CS, text="Inserte nombre de la hoja:")
	excel_stocks_hoja_lab.grid(column=1, row=3)
	global excel_stocks_hoja
	excel_stocks_hoja = Entry(frame_CS, width=15)
	excel_stocks_hoja.grid(column=2, row=3, padx=5, pady=5)
	excel_stocks_columna_isbn_lab= Label(frame_CS, text="Inserte el número de columna de ISBN:")
	excel_stocks_columna_isbn_lab.grid(column=1, row=4)
	global excel_stocks_columna_isbn_ent
	excel_stocks_columna_isbn_ent= Entry(frame_CS, width=5)
	excel_stocks_columna_isbn_ent.grid(column=2, row=4, padx=5, pady=5)
	excel_stocks_columna_stocks_lab= Label(frame_CS, text="Inserte el número de columna de stock:")
	excel_stocks_columna_stocks_lab.grid(column=1, row=5)
	global excel_stocks_columna_stocks_ent
	excel_stocks_columna_stocks_ent= Entry(frame_CS, width=5)
	excel_stocks_columna_stocks_ent.grid(column=2, row=5, padx=5, pady=5)

	actualizar_stocks = Button(frame_CS2, text="Controlar stocks", command=iniciador)
	actualizar_stocks.grid(column=2, row=4, padx=5, pady=5)
	global info2
	info2 = Frame(frame_CSS)
	info2.pack()
	inlabel = Label(info2, text="Registros que serán procesados: ")
	inlabel.grid(column=1, row=2)
	global inlabelpro2
	inlabelpro2 = Label(info2, text="0")
	inlabelpro2.grid(column=2, row=2)
	botonera2 = Frame(frame_CSS)
	botonera2.pack()

	boton_de_error_de_vacio = Button(botonera2, text="Stocks en 0 en Ultra", command=desplegadordevacio_stock)
	boton_de_error_de_vacio.grid(column=1, row=4, padx=5, pady=5)
	boton_de_error_de_API = Button(botonera2, text="Errores de API", command=desplegadordeAPI_stock)
	boton_de_error_de_API.grid(column=2, row=4, padx=5, pady=5)
	boton_de_cambiados = Button(botonera2, text="Publicaciones Cambiadas", command=desplegadorcambiados_stock)
	boton_de_cambiados.grid(column=3, row=4, padx=5, pady=5)
	boton_stockbajos =Button(botonera2, text="Control de Stock en 0", command=desplegarstockbajos)
	boton_stockbajos.grid(column=1, row=3, padx=5, pady=5)
	boton_stockaltos =Button(botonera2, text="Control de ML en 0", command=desplegarstockaltos)
	boton_stockaltos.grid(column=2, row=3, padx=5, pady=5)
	boton_stockaltospau =Button(botonera2, text="Control de ML pausados", command=desplegastocksaltospausados)
	boton_stockaltospau.grid(column=3, row=3, padx=5, pady=5)
#Cambiar Usuario de ML	
def cambiarusuarioguardar():
	global usuario
	global entry_usuario
	usuario = entry_usuario.get()
	if usuario == str(553279780):
		nombre_de_usuario_label.config(text="Editorial Losada")
	elif usuario == str(439527517):
		nombre_de_usuario_label.config(text="AIQUE grupo editorial")
	entry_usuario.destroy()
	numero_de_usuario_label.config(text=usuario)
	boton_cambio_de_usuario.config(text="Cambiar usuario", command=cambiarusuario)

def cambiarusuario():
	global entry_usuario
	entry_usuario = Entry(frame_US, width=8)
	entry_usuario.grid(column=1, row=1,padx=10, pady=10)
	nombre_de_usuario_label.config(text="")
	boton_cambio_de_usuario.config(text="Guardar usuario", command=cambiarusuarioguardar)	

#Corregir títulos
def titulador(mlitem, token):
	url2 = "https://api.mercadolibre.com/items/" + str(mlitem)
	headers2 = {'Authorization' : ('Bearer '+ token), 'Content-type': 'application/json', 'Accept':'application/json'}
	payload = {'title': dic_nuevotitulo[mlitem]}
	response2 = requests.put(url2, headers=headers2, json=payload)
	print(response2)
	if response2.status_code == 200:
		lis_cambiados_titulo.append(mlitem)
	else:
		lis_erroresdeapi_titulo.append(mlitem)	
def funerariatitulo(ml):
	try:
		print(str(ml)+dic_titulo_titulo[ml])
		ltit = dic_titulo_titulo[ml].split(',')
		lautor = dic_autor_titulo[ml].split(',')
		lautorx= lautor[0].strip(" ")
		ltitcentro = ltit[0].strip(" ")
		ltitart = ltit[1]
		if ltitart[0] == " ":
			if ltitart[3] == " ":
				ltitarticulo= ltitart[1]+ltitart[2]
			else:	
				ltitarticulo= ltitart[1]+ltitart[2]+ltitart[3]
		else:
			if ltitart[2] == " ":
				ltitarticulo= ltitart[0]+ltitart[1]
			else:	
				ltitarticulo= ltitart[0]+ltitart[1]+ltitart[2]
		if len(ltit)>1:
			dic_nuevotitulo[ml]=str(ltitarticulo+" "+ltitcentro+" - "+lautorx+" - "+dic_editorial[ml])
		else:
			dic_nuevotitulo[ml]=str(ltit[0]+" - "+lautor[0]+" - "+dic_editorial[ml])
	except IndexError:
		print("chao")		
		print(str(ml)+dic_titulo_titulo[ml])
		ltit = dic_titulo_titulo[ml].split(',')
		lautor = dic_autor_titulo[ml].split(',')
		lautorx= lautor[0].strip(" ")
		ltitcentro = ltit[0].strip(" ")
		dic_nuevotitulo[ml]=str(ltitcentro+" - "+lautorx+" - "+dic_editorial[ml])
	f = (len(dic_nuevotitulo)+1)
	print("CORTADO:")
	tra = dic_nuevotitulo[ml]
	trabajando=tra[0:59]
	dic_nuevotitulo[ml]=trabajando
	print(dic_nuevotitulo[ml])	

def isbeniador_titulo(ml, token, hoja_titulo, columna_isbn, columa_titulo):
	try:
		fa = (len(hoja_titulo['A'])+1)
		for y in range(2,fa):
			isbnm = dic_isbn_titulo[ml]
			isbnu = hoja_titulo.cell(row=y, column=int(columna_isbn)).value
			if str(isbnm) == str(isbnu):
				tit = hoja_titulo.cell(row=y, column=2).value
				dic_titulo_titulo[ml]= str(tit)
				aut = hoja_titulo.cell(row=y, column=3).value
				dic_autor_titulo[ml]= str(aut)
				edi = hoja_titulo.cell(row=y, column=4).value
				dic_editorial[ml] = str(edi)
		funerariatitulo(ml)		
		titulador(ml, token)
	except KeyError:
		lis_erroresdevacio_titulo.append(ml)
def diccionador_titulo(libro2, hoja2, columna_isbn, columa_titulo):
	lista_titulo = op.load_workbook(libro2 +'.xlsx')
	hoja_titulo = lista_titulo[hoja2]
	#token
	token = tok
	global revisor_desdeExcel
	global revisor_desdeAPI
	if revisor_desdeExcel == True:
		global libro_publica
		libro1 = libro_publica
		global hoja_publica
		hoja1 = hoja_publica
		lista_publicaciones = op.load_workbook(libro1 +'.xlsx')
		hoja_publicaciones = lista_publicaciones[hoja1]
		ga = (len(hoja_publicaciones['A'])+1)
		for x in range (2,ga):
			global columna_ml
			item = hoja_publicaciones.cell(row=x, column=int(columna_ml)).value 
			global columna_isbn_publi
			isbn = hoja_publicaciones.cell(row=x, column=int(columna_isbn_publi)).value
			if isbn != None:
				lista_ML_titulo.append(item)
				dic_isbn_titulo[item]=isbn
				url = "https://api.mercadolibre.com/items/" + str(item)
				headers = {'Authorization': ('Bearer '+ token), 'Content-type': 'application/json', 'Accept':'application/json'}
				response = requests.get(url, headers=headers)
				if response.status_code == 200:
					jsonp = response.json()
					titul=jsonp['title']
					dic_titulo_titulo[item]=titul
			lis_procesados_titulo.append(item)		
	# titulos del ultra
			isbeniador_titulo(item, token, hoja_titulo, columna_isbn, columa_titulo)
	else:
		messagebox.showerror(message="No está cargado el catálogo")				
	
	if len(lis_fuera_titulo) > 0:
		messagebox.showerror(message=str(len(lis_fuera_titulo)) + " registros, no se encuentran en la base de datos.", title="Error de Api" )
	if len(lis_erroresdevacio_titulo) > 0:
		messagebox.showerror(message=str(len(lis_erroresdevacio_titulo)) + "No se pudo actualizar por estar vacíos", title="Error" )
	messagebox.showinfo(message=str(len(lis_cambiados_titulo))+ "fueron cambiados de titulo", title="Acualización de titulo")
def botonactualizar_titulo():
	excel_titulo_libro
	libro2 = excel_titulo_libro.get()
	hoja2 = excel_titulo_hoja.get()
	columna_isbn=excel_titulo_columna_isbn_ent.get()
	columa_titulo=excel_titulo_columna_titulo_ent.get()
	diccionador_titulo(libro2, hoja2, columna_isbn, columa_titulo)
def cerraractualizadordetitulo():
	bot_actualizar_titulo.config(command=actualizador_de_titulo)
	global frame_AP
	frame_AP.destroy()
def actualizador_de_titulo():
	bot_actualizar_titulo.config(command=cerraractualizadordetitulo)
	global frame_AP
	frame_AP = Frame(frame_activo)
	frame_AP.pack()
	frame5= Frame(frame_AP)
	frame5.pack()
	frame6= Frame(frame_AP)
	frame6.pack()
	#Ingresar lista de titulo
	excel_titulo_libro_lab = Label(frame5, text="Inserte Libro de titulo:")
	excel_titulo_libro_lab.grid(column=1, row=2)
	excel_titulo_libro_lab2 = Label(frame5, text=".xlsx")
	excel_titulo_libro_lab2.grid(column=3, row=2)
	global excel_titulo_libro
	excel_titulo_libro = Entry(frame5, width=15)
	excel_titulo_libro.grid(column=2, row=2, padx=5, pady=5)
	excel_titulo_hoja_lab = Label(frame5, text="Inserte nombre de la hoja:")
	excel_titulo_hoja_lab.grid(column=1, row=3)
	global excel_titulo_hoja
	excel_titulo_hoja = Entry(frame5, width=15)
	excel_titulo_hoja.grid(column=2, row=3, padx=5, pady=5)
	excel_titulo_columna_isbn_lab= Label(frame5, text="Inserte el número de columna de ISBN:")
	excel_titulo_columna_isbn_lab.grid(column=1, row=4)
	global excel_titulo_columna_isbn_ent
	excel_titulo_columna_isbn_ent= Entry(frame5, width=5)
	excel_titulo_columna_isbn_ent.grid(column=2, row=4, padx=5, pady=5)
	excel_titulo_columna_titulo_lab= Label(frame5, text="Inserte el número de columna de titulo:")
	excel_titulo_columna_titulo_lab.grid(column=1, row=5)
	global excel_titulo_columna_titulo_ent
	excel_titulo_columna_titulo_ent= Entry(frame5, width=5)
	excel_titulo_columna_titulo_ent.grid(column=2, row=5, padx=5, pady=5)

	actualizar_titulo = Button(frame6, text="Actualizar", command=botonactualizar_titulo)
	actualizar_titulo.grid(column=2, row=4, padx=5, pady=5)
	global info
	info = Frame(frame_AP)
	info.pack()
	inlabel = Label(info, text="Registros que serán procesados: ")
	inlabel.grid(column=1, row=2)
	global inlabelpro
	inlabelpro = Label(info, text="0")
	inlabelpro.grid(column=2, row=2)
	botonera = Frame(frame_AP)
	botonera.pack()

#RESPONDEDOR
stock="Tenemos stock de este libro."
lugar="Nos encontramos en San Nicolás en CABA, cerca del teatro San Martín."
horario="Nuestro horario de atención en la librería es de 11:00 a 18:00 de lunes a sábados"
enviopais="Hacemos envíos a todo el país."
envioamba="Si estás en el AMBA podés utilizar nuestro servicio Flex de envíos en el día, hacemos entregas en moto de 15:00 a 20:00."
retiro="Si lo compras con acordar para el vendedor lo podés retirar inmediatamente por nuestro local."
conjunto="Podés llevar varios productos en la misma compra, con el mismo valor de envío. Para esto debes agregarlos al carrito, con el botón que está debajo del botón comprar. Una agregados todos los productos, podes comprar el carrito, entrándolo en el ícono en la esquina superior derecha de la pantalla."
enviogratis="Para algunas zonas del país por compras superiores a $3500 el envío es gratis."
pregunta_activa=""

global c_stock
c_stock = False
clugar=False
siguientepregunta=1
def responder():
	token=tok
	global respuesta_text
	respuesta = respuesta_text.get(1.0,END)
	data={
	'question_id':pregunta_activa,
	'text':respuesta
	}
	url= "https://api.mercadolibre.com/answers"
	headers = {'Authorization': ('Bearer '+ token), 'Content-type': 'application/json', 'Accept':'application/json'}
	response = requests.post(url, headers=headers, json=data)
	conte = response.content
	if response.status_code == 200:
		messagebox.showinfo(message="Pregunta respondida")
		respuesta_text.delete(1.0,END)	
		respuesta_text.insert(END,"Hola!Gracias por tu consulta. Librería Losada")
	else:
		messagebox.showerror(message="Error respondiendo pregunta")	
	descargarPreguntas()
			
def desplegar_preguntas(n):
	preg =lista_preguntas[n]
	global pregunta_activa
	pregunta_activa = preg['id']
	global pregunta
	pregunta.config(text=preg['text'])
	token=tok
	url= "https://api.mercadolibre.com/items/"+preg['item_id']
	headers = {'Authorization': ('Bearer '+ token), 'Content-type': 'application/json', 'Accept':'application/json'}
	response = requests.get(url, headers=headers)
	if response.status_code == 200:
		jsonp = response.json()
		titulo = jsonp['title']
		atributos = jsonp['attributes']
		at=(len(atributos)-1)
		atributos_sku = atributos[at]
		strisbn = atributos_sku['value_name']
		try:
			isbn=int(strisbn)
		except ValueError:
			isbn=1234567891011
		global libro_titulo	
		libro_titulo.config(text=titulo)
		global isbn_titulo
		isbn_titulo.config(text=isbn)

def descargarPreguntas():
	global siguientepregunta
	siguientepregunta = 1
	token=tok
	url= "https://api.mercadolibre.com/my/received_questions/search?status=UNANSWERED"
	headers = {'Authorization': ('Bearer '+ token), 'Content-type': 'application/json', 'Accept':'application/json'}
	response = requests.get(url, headers=headers)
	if response.status_code == 200:
		jsonp = response.json()
		global lista_preguntas
		lista_preguntas=jsonp['questions']
		global num_por_responder_label
		num_por_responder_label.config(text=(len(lista_preguntas)))
		if len(lista_preguntas) > 0:
			desplegar_preguntas(0)
		else:
			global frame_activo_respondedor
			frame_activo_respondedor.destroy()

class Cambiar:
	def cambios(item, vara, control):
		global respuesta_text
		esta=respuesta_text.search(item,1.0,END)
		if esta == "":
			control=False
		else:
			control=True	
		if control == False:
			var = vara.get()
			if var == 1:
				respuesta_text.insert(1.5,item)
				control=True
		elif control == True:
			lug= respuesta_text.search(item,1.0,END)
			x=len(item)
			term = item[-3:x]
			prefin =respuesta_text.search(term,1.0,END)
			pprefin=prefin.split(".")
			fin = str(pprefin[0])+"."+str((int(pprefin[1])+3))
			respuesta_text.delete(lug ,fin)
			control=False	

def cambio_stock():
	global var_stock
	Cambiar.cambios(stock,var_stock,c_stock)
def cambio_lugar():
	global var_lugar
	Cambiar.cambios(lugar,var_lugar,clugar)
def cambio_horario():
	global var_horario
	Cambiar.cambios(horario,var_horario,clugar)
def cambio_enviopais():
	global var_enviopais
	Cambiar.cambios(enviopais,var_enviopais,clugar)
def cambio_envioamba():
	global var_envioamba
	Cambiar.cambios(envioamba,var_envioamba,clugar)
def cambio_retiro():
	global var_retiro
	Cambiar.cambios(retiro,var_retiro,clugar)
def cambio_conjunto():
	global var_conjunto
	Cambiar.cambios(conjunto,var_conjunto,clugar)
def cambio_enviogratis():
	global var_enviogratis
	Cambiar.cambios(enviogratis,var_enviogratis,clugar)			

def saltarpregunta():
	global siguientepregunta
	if len(lista_preguntas) > 0:
		desplegar_preguntas(siguientepregunta)
		siguientepregunta+=1
	else:
		global frame_activo_respondedor
		frame_activo_respondedor.destroy()
#TOPLEVER DE RESPONDEDOR
def respondedor():
	respondor_toplevel = Toplevel(raiz)
	respondor_toplevel.title("Respondedor - API - Losada")
	frame0_responder = Frame(respondor_toplevel)
	frame0_responder.grid(column=0, row=1, padx=5, pady=5)
	global frame_activo_respondedor
	frame_activo_respondedor = Frame(respondor_toplevel)
	frame_activo_respondedor.grid(column=0, row=2, padx=5, pady=5)
	frame_respuesta = Frame(respondor_toplevel)
	frame_respuesta.grid(column=0, row=3, padx=5, pady=5)
	frame_checkbuttons= Frame(frame_respuesta)
	frame_checkbuttons.grid(column=2, row=1, padx=5, pady=5)

	global pregunta
	pregunta = Label(frame_activo_respondedor,text="")
	pregunta.grid(column=1, row=2, padx=5, pady=5)
	global libro_titulo	
	libro_titulo=Label(frame_activo_respondedor,text="")
	libro_titulo.grid(column=1, row=1, padx=5, pady=5)
	global isbn_titulo
	isbn_titulo=Label(frame_activo_respondedor,text="")
	isbn_titulo.grid(column=2, row=1, padx=5, pady=5)

	bot_descargar_preguntas = Button(frame0_responder, text="Actualizar Preguntas", command=descargarPreguntas)
	bot_descargar_preguntas.grid(column=1, row=1, padx=5, pady=5)
	pre_por_responder_label = Label(frame0_responder, text="Preguntas por responder:")
	pre_por_responder_label.grid(column=2, row=1 , pady=5)
	global num_por_responder_label
	num_por_responder_label = Label(frame0_responder, text=0)
	num_por_responder_label.grid(column=3, row=1, pady=5)

	global respuesta_text
	respuesta_text = Text(frame_respuesta,width=50,height=7)
	respuesta_text.insert(END, "Hola!Gracias por tu consulta. Librería Losada")
	respuesta_text.grid(column=1, row=1, padx=5, pady=5)
	global var_stock
	var_stock = IntVar()
	Checkbutton(frame_checkbuttons, text="stock", variable=var_stock, command=cambio_stock).grid(row=1, column=1)
	#list_var_stock.append(var_stock)
	global var_lugar
	var_lugar = IntVar()
	Checkbutton(frame_checkbuttons, text="lugar", variable=var_lugar,command=cambio_lugar).grid(row=2, column=1)
	#list_var_lugar.append(var_lugar)
	global var_horario
	var_horario = IntVar()
	Checkbutton(frame_checkbuttons, text="horario", variable=var_horario,command=cambio_horario).grid(row=3, column=1)
	#list_var_horario.append(var_horario)
	global var_enviopais
	var_enviopais = IntVar()
	Checkbutton(frame_checkbuttons, text="envio PAIS", variable=var_enviopais,command=cambio_enviopais).grid(row=4, column=1)
	#list_var_envio.append(var_envio)
	global var_envioamba
	var_envioamba = IntVar()
	Checkbutton(frame_checkbuttons, text="envio AMBA", variable=var_envioamba,command=cambio_envioamba).grid(row=5, column=1)
	global var_retiro
	var_retiro = IntVar()
	Checkbutton(frame_checkbuttons, text="retiro", variable=var_retiro,command=cambio_retiro).grid(row=6, column=1)
	global var_conjunto
	var_conjunto = IntVar()
	Checkbutton(frame_checkbuttons, text="conjunto", variable=var_conjunto,command=cambio_conjunto).grid(row=1, column=2)
	global var_enviogratis
	var_enviogratis = IntVar()
	Checkbutton(frame_checkbuttons, text="envío gratis", variable=var_enviogratis,command=cambio_enviogratis).grid(row=2, column=2)
	bot_responder = Button(frame_checkbuttons, text="Responder", command=responder)
	bot_responder.grid(column=1, row=7, padx=5, pady=5)
	bot_saltar = Button(frame_checkbuttons, text="Saltar Pregunta", command=saltarpregunta)
	bot_saltar.grid(column=1, row=8, padx=5, pady=5)	
#POSTEAR PUBLICACIONES

def formato_correcto():
	mb.showinfo(
		'Formato correcto',
		'Los URLs deben ser ingresados en una línea separados por espacios.'
		+ '\n con la forma: http://***/ISBN001.jpg \n' +
		'Donde 001.jpg será la portada, 002.jpg será la contraportada \n'
		+ 'se puede incluir hasta 006.jpg.\n' +
		'El concatenador soporta ISBN 10 y EAN13'
		)

def excluidos():
	nv = Toplevel(raiz)
	cuadro_excluidos = Text(nv, width=75, height=25)
	cuadro_excluidos.pack()
	for isbndn in no_en_la_base:
		cuadro_excluidos.insert(END, isbndn + "," + imagenes[isbndn] + "\n") 

def publicador():
	base_publicador = Toplevel(raiz)
	base_publicador.title("Librería Losada")
	frame_publicador = Frame(base_publicador, width=1200, height=500)
	frame_publicador.pack()
	bienvenida = Label(frame_publicador, text="Concatenador de imágenes")
	bienvenida.grid(column=1, row=0)
	ingrese = Label(frame_publicador, text="Ingrese URL de imágenes:")
	ingrese.grid(column=0, row=1)

	ingreso = Entry(frame_publicador, width=75)
	ingreso.grid(column=1, row=1, padx=5, pady=5)
	boton_contatenador = Button(frame_publicador, text="Concatenar URL", width=15, height=5, command=concatenar)
	boton_contatenador.grid(column=3, row=3,padx=5, pady=5)

	resultado = Label(frame_publicador, text="URL concatenados:")
	resultado.grid(column=0, row=2, padx=5, pady=5),

	cuadro_resultado = Text(frame_publicador, width=75, height=25)
	cuadro_resultado.grid(column=1, row=3, padx=5, pady=5)

	scroll_publicador = Scrollbar(frame_publicador, command=cuadro_resultado.yview)
	scroll_publicador.grid(column=2, row=3, sticky="nsew")
	cuadro_resultado.config(yscrollcommand=scroll_publicador.set)

	boton_de_formato = Button(frame_publicador, text="Ver formato correcto", command=formato_correcto)
	boton_de_formato.grid(column=3, row=1,padx=5, pady=5)

	boton_de_formato = Button(frame_publicador, text="Ver URLs excluidos", command=excluidos)
	boton_de_formato.grid(column=3, row=4,padx=5, pady=5)

#GUI - Base
raiz = Tk()
raiz.title("Librería Losada - API Mercado Libre")
frame_US = Frame(raiz)
frame_US.pack()
frame0 = Frame(raiz)
frame0.pack()
frame_activo = Frame(raiz)
frame_activo.pack()

frame1 = Frame(frame0)
frame1.grid(column=1, row=2,padx=10, pady=10)
frame2 = Frame(frame0)
frame2.grid(column=2, row=2,padx=10, pady=10)

#Frame1- Status
st_label_titulo = Label(frame1, text="Status:")
st_label_titulo.grid(column=1, row=1, padx=5, pady=5)
st_label_ttoken = Label(frame1, text="Token:")
st_label_ttoken.grid(column=1, row=2, padx=5, pady=5)
st_label_rtoken = Label(frame1, text="No cargado")
st_label_rtoken.grid(column=2, row=2, padx=5, pady=5)
st_label_tcatalogo = Label(frame1, text="Catálogo:")
st_label_tcatalogo.grid(column=1, row=3, padx=5, pady=5)
st_label_rcatalogo = Label(frame1, text="No cargado")
st_label_rcatalogo.grid(column=2, row=3, padx=5, pady=5)

#Frame2 - Botones
bot_obtener_token = Button(frame2, text="Obtener Token", command=obtenertoken)
bot_obtener_token.grid(column=1, row=1, padx=5, pady=5)
bot_cargar_catalogo = Button(frame2, text="Cargar Catálogo", command=cargarcatalogo)
bot_cargar_catalogo.grid(column=2, row=1, padx=5, pady=5)
bot_actualizar_precios = Button(frame2, text="Actualizar Precios", command=actualizador_de_precios)
bot_actualizar_precios.grid(column=1, row=2, padx=5, pady=5)
bot_controlar_stock = Button(frame2, text="Controlar Stock", command=controlstocks)
bot_controlar_stock.grid(column=2, row=2, padx=5, pady=5)
bot_actualizar_titulo = Button(frame2, text="Actualizar Títulos", command=actualizador_de_titulo)
bot_actualizar_titulo.grid(column=1, row=3, padx=5, pady=5)
bot_respondedor = Button(frame2, text="Respondedor", command=respondedor)
bot_respondedor.grid(column=2, row=3, padx=5, pady=5)
bot_publicador = Button(frame2, text="Publicar", command=publicador)
bot_publicador.grid(column=2, row=4, padx=5, pady=5)

#Selección de vendedor de ML
usu_lab = Label(frame_US, text="Usuario:")
usu_lab.grid(column=1, row=1,padx=10, pady=10)
numero_de_usuario_label = Label(frame_US, text="553279780")
numero_de_usuario_label.grid(column=2, row=1,padx=10, pady=10)
nombre_de_usuario_label = Label(frame_US, text="Editorial Losada")
nombre_de_usuario_label.grid(column=3, row=1,padx=10, pady=10)
boton_cambio_de_usuario = Button(frame_US, text="Cambiar usuario", command=cambiarusuario)
boton_cambio_de_usuario.grid(column=4, row=1,padx=10, pady=10)
raiz.mainloop()	
	