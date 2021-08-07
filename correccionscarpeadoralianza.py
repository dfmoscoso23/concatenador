#correcciónscrapper
import pandas as pd
import openpyxl as op
import os
import requests
from bs4 import BeautifulSoup


excel = op.load_workbook("listaAlianza_1320.xlsx")
hoja=excel.active
"""
for x in range(2,42):
	isbn=str(hoja.cell(row=x,column=3).value)
	print(isbn)
	url="https://www.alianzaeditorial.es/busqueda.php?tipobusqueda=busqueda&precioMin=0&precioMax=200&texto="+isbn
	r = requests.get(url)
	soup = BeautifulSoup(r.text, 'lxml')
	busca=soup.find('div', class_="book-list-item-image")
	href=busca.a['href']
	urlo='https://www.alianzaeditorial.es'+href
	ro = requests.get(urlo)
	soup = BeautifulSoup(ro.text, 'lxml')
	precio=soup.find('option', class_="btn-formato")
	hoja.cell(row=x,column=6).value=precio.text
	excel.save('C:/Users/David/Desktop/listaAlianza.xlsx')
"""

eml = pd.read_csv(r"C:/Users/David/Desktop/EML.txt", sep=';')
i=len(hoja['A'])+1
for x in range(2,i):
	isbn_box=hoja.cell(row=x,column=3).value
	isbn_deco=isbn_box.split("-")
	isbn=str(isbn_deco[0])+str(isbn_deco[1])+str(isbn_deco[2])+str(isbn_deco[3])+str(isbn_deco[4])
	fila= eml[eml.ISBN==isbn]
	if fila.empty:
		hoja.cell(row=x,column=5).value=0
	else:
		hoja.cell(row=x,column=5).value=fila.iat[0,17]
	excel.save('C:/Users/David/Desktop/listaAlianza_1320.xlsx')	
	