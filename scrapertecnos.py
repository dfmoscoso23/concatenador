#scrapeador tecnos
import pandas as pd
import openpyxl as op
import os
import requests
from bs4 import BeautifulSoup

nombre="Tecnos"

lista_isbn=[]

excel = op.load_workbook("listaTecnos.xlsx")
hoja=excel.active

fin = False

eml = pd.read_csv(r"C:/Users/David/Desktop/EML.txt", sep=';')
x=len(hoja['A'])
while fin == False:
	url=input("Ingrese URL:")
	if url == "Z":
		fin=True
	else:
		x=(len(hoja['A'])+1)	
		r = requests.get(url)
		soup = BeautifulSoup(r.text, 'lxml')
		#primer_grupo=soup.find('div', class_="ficha-contenidos")
		info_basica=soup.find('div', class_="ficha-contenidos")
		hoja.cell(row=x,column=1).value=info_basica.h1.text
		hoja.cell(row=x,column=2).value=info_basica.a.text
		info_detallada=soup.find('div', class_="datos")
		datali=info_detallada.find_all('b')
		isbn_box=datali[4].next_sibling
		hoja.cell(row=x,column=3).value=isbn_box
		precio=datali[3].next_sibling
		hoja.cell(row=x,column=6).value=precio
		isbn_deco=isbn_box.split("-")
		isbn=str(isbn_deco[0].strip())+str(isbn_deco[1])+str(isbn_deco[2])+str(isbn_deco[3])+str(isbn_deco[4])
		print(isbn)
		print(info_basica.h1.text)
		print(info_basica.a.text)
		fila= eml[eml.ISBN==isbn]
		if fila.empty:
			hoja.cell(row=x,column=5).value=0
		else:
			hoja.cell(row=x,column=5).value=fila.iat[0,17]
		hoja.cell(row=x,column=4).value="Tecnos editorial"
		excel.save('C:/Users/David/Desktop/lista'+nombre+'.xlsx')
			
		
	
"""		
x=1
for isbn in lista_isbn:
	fila= eml[eml.ISBN==isbn]
	hoja.cell(row=x,column=1).value=isbn
	hoja.cell(row=x,column=2).value=fila.iat[0,1]
	hoja.cell(row=x,column=3).value=fila.iat[0,2]
	hoja.cell(row=x,column=4).value="Obras Completas"
	hoja.cell(row=x,column=5).value=fila.iat[0,28]
	try:
		raw_isbn=isbn
		isbn_guinado=raw_isbn[0:3]+"-"+raw_isbn[3:6]+"-"+raw_isbn[6:8]+"-"+raw_isbn[8:12]+"-"+raw_isbn[-1]
		url='http://www.editoriallosada.com/search/content/'+isbn_guinado
		r = requests.get(url)
		soup = BeautifulSoup(r.text, 'lxml')
		busca=soup.find('div', class_="ds-2col-fluid node node-libro node-teaser view-mode-teaser clearfix")
		href=busca.h2.a['href']
		urlo='http://www.editoriallosada.com'+href
		ro = requests.get(urlo)
		soupo = BeautifulSoup(ro.text, 'lxml')
		div_paginas=soupo.find('div', class_="field field-name-field-paginas field-type-text field-label-inline clearfix")
		paginas=div_paginas.find('div', class_="field-item even")
		sinopsis=soupo.find('div', class_="lead")
		div_formato=soupo.find('div', class_="field field-name-field-formato field-type-text field-label-hidden")
		formato=div_formato.find('div', class_="field-item even").text
		if "cm" in formato:
			forma=formato.split("cm")
			formato3=forma[0]
			formato2=formato3.split("x")
			if int(formato2[0].strip()) > int(formato2[1].strip()):
				alto=formato2[0]
				ancho=formato2[1]
			else:
				alto=formato2[1]
				ancho==formato2[0]		
		hoja.cell(row=x,column=6).value=paginas.text
		hoja.cell(row=x,column=7).value=alto
		hoja.cell(row=x,column=8).value=ancho
		hoja.cell(row=x,column=9).value=sinopsis.text
	except AttributeError:
		print("no en la página web")
	x+=1

	print(fila.iat[0,2])
	#print(fila['Autor'])
excel.save('C:/Users/David/Desktop/lista'+nombre+'.xlsx')
"""