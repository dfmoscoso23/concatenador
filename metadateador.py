#metadateador Losada
import pandas as pd
import openpyxl as op
import os
import requests
import re
from bs4 import BeautifulSoup
from shutil import copyfile

nombre="OMP"

lista_isbn=[]

excel = op.Workbook()
hoja=excel.active

def cambiarcomaporpunto(entrada):
	result=re.sub(r",",".",entrada)
	return result

eml = pd.read_csv(r"C:/Users/David/Desktop/EML.txt", sep=';',error_bad_lines=False)

for root, dirs, files in os.walk(".", topdown=False):
	for name in files:
		if name[-5:-1]=="1.jp":
			isbn_precoz = name.split("001.")
			lista_isbn.append(isbn_precoz[0])
			a_copiar="C:/Users/David/Desktop/librosya/"+isbn_precoz[0]+".jpg"
			copyfile(name, a_copiar)
print(lista_isbn)
x=1
for isbn in lista_isbn:
	try:
		fila= eml[eml.ISBN==isbn]
		hoja.cell(row=x,column=1).value=isbn
		hoja.cell(row=x,column=2).value=fila.iat[0,1]
		hoja.cell(row=x,column=3).value=fila.iat[0,2]
		hoja.cell(row=x,column=4).value="Grandes Clásiscos"
		hoja.cell(row=x,column=5).value=fila.iat[0,28]
	except IndexError:
		print(str(isbn)+"no en ULTRA")	
		pass
	try:
		raw_isbn=isbn
		isbn_guinado=raw_isbn[0:3]+"-"+raw_isbn[3:6]+"-"+raw_isbn[6:8]+"-"+raw_isbn[8:12]+"-"+raw_isbn[-1]
		url='http://www.editoriallosada.com/search/content/'+isbn_guinado
		r = requests.get(url)
		soup = BeautifulSoup(r.text, 'lxml')
		busca=soup.find('div', class_="ds-2col-fluid node node-libro node-teaser view-mode-teaser clearfix")
		href=busca.h2.a['href']
		urlo='http://www.editoriallosada.com'+href
		ro = requests.get(urlo)
		soupo = BeautifulSoup(ro.text, 'lxml')
		div_paginas=soupo.find('div', class_="field field-name-field-paginas field-type-text field-label-inline clearfix")
		paginas=div_paginas.find('div', class_="field-item even")
		sinopsis=soupo.find('div', class_="lead")
		div_formato=soupo.find('div', class_="field field-name-field-formato field-type-text field-label-hidden")
		formato=div_formato.find('div', class_="field-item even").text
		if "," in formato:
			formato=cambiarcomaporpunto(formato)
		if "cm" in formato:
			forma=formato.split("cm")
			formato3=forma[0]
			formato2=formato3.split("x")
			if float(formato2[0].strip()) > float(formato2[1].strip()):
				alto=formato2[0]
				ancho=formato2[1]
			else:
				alto=formato2[1]
				ancho=formato2[0]
		else:
			formato2=formato.split("x")
			if float(formato2[0].strip()) > float(formato2[1].strip()):
				alto=formato2[0]
				ancho=formato2[1]
			else:
				alto=formato2[1]
				ancho=formato2[0]				
		hoja.cell(row=x,column=6).value=paginas.text
		hoja.cell(row=x,column=7).value=alto
		hoja.cell(row=x,column=8).value=ancho
		hoja.cell(row=x,column=9).value=sinopsis.text
	except AttributeError:
		print("no en la página web")
	x+=1
	try:
		print(fila.iat[0,2])
	except IndexError:
		pass	
	#print(fila['Autor'])
excel.save('C:/Users/David/Desktop/lista'+nombre+'.xlsx')