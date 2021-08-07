#Backstage
import openpyxl as op
import pandas as pd
#Diccionarios
dic_titulo = {}
dic_autor = {}
dic_editorial = {}
dic_precio = {}
dic_titulopub = {}
portada = {}
contraportada = {}
paginaextra1 = {}
paginaextra2 = {}
paginaextra3 = {}
paginaextra4 = {}
dic_tema = {}
apellido = {}
imagenes = {}
isbns = []
isbn_depurado = []
en_la_base = []
no_en_la_base = []
global excel
excel=False
global csvcon
csvcon=False

def funerariat(tit, isbnd):
		if ',' in tit:
			titu = tit.split(',')
			arti = titu [1].strip(" ")
			titul = titu[0].strip(" ")
			try:
				if arti[0]==" ":
					if arti[3]==" ":
						artic= arti[1]+arti[2]
					elif arti[3]=="s" or arti[3]=="o" or arti[3]=="a":
						artic= arti[1]+arti[2]+arti[3]
				else:
					if arti[2]==" ":
						artic=	arti[0]+arti[1]
					elif arti[2]=="s" or arti[2]=="o" or arti[2]=="a":
						artic= arti[0]+arti[1]+arti[2]
					else:
						artic=""	
				titulo = artic + " " + titul
			except IndexError:
				titulo = tit	
		else:
			titulo = tit
		tituloo = titulo.title()
		titulod = tituloo.strip()
		dic_titulo[isbnd] = titulod	
def funerariaa(aut, isbnd):
	if ',' in aut:
		auto = aut.split(',')
		artic = auto [1].strip(" ")
		ape = auto[0].strip(" ")
		autor = artic + " " + ape
		apellido[isbnd] = ape.title()
	else:
		autor = aut.strip("	")
		apellido[isbnd] = autor.title()
	autort = autor.strip(" ")
	autord = autort.title()
	dic_autor[isbnd] = autord


def funerariae(edit, isbnd):
	editora = edit.title()
	editorial = editora.strip(" ")
	dic_editorial[isbnd] = editorial
def funerariap(pre, isbnd):
	dic_precio[isbnd] =str(pre)	

def concatenadopub(isbnd):
	dic_titulopub[isbnd] = dic_titulo[isbnd] + " - "+ apellido[isbnd] + " - " + dic_editorial[isbnd]

def funerariatema(isbnd, tem):
	dic_tema[isbnd] = tem.title()

def busqueda(xrow, isbnd, x):
	#cuadro_resultado.insert(END, isbnd)
	#cuadro_resultado.insert(END, xrow)
	if isbnd==str(xrow):
		en_la_base.append(isbnd)
		global sheet	
		#print(sheet.cell(row=x, column=2).value)
		tit = sheet.cell(row=x, column=2).value
		aut = sheet.cell(row=x, column=3).value
		edit = sheet.cell(row=x, column=4).value
		pre = sheet.cell(row=x, column=11).value
		tem = sheet.cell(row=x, column=6).value
		funerariat(tit, isbnd)
		funerariaa(aut, isbnd)
		funerariae(edit, isbnd)
		funerariap(pre, isbnd)
		funerariatema(isbnd, tem)
		concatenadopub(isbnd)
		cuadro_resultado.insert(
			END, dic_titulopub[isbnd] + "," +
			isbnd + "," + imagenes[isbnd] + "," +
			isbnd + "," + "1" + "," + dic_precio[isbnd] + "," + "Nuevo" + "," + "des" + "," + " " + "," +
			"Clásica" + "," + "Mercado Envíos | Mercado Envíos Flex" + "," + "A cargo del comprador" + "," +
			"Acepto" + "," + "Garantía del vendedor"  + "," + "1" + "," + "meses"  + "," + "Papel" + "," + 
			dic_tema[isbnd]  + "," + dic_titulo[isbnd] + "," + dic_autor[isbnd] + "," + "Español" + "," + 
			dic_editorial[isbnd] + "," + dic_tema[isbnd] + "\n"
			)
	else:
		no_en_la_base.append(isbnd)	
def busqudacsv(isbnd):
	global csv
	fila_isbn=csv[csv.ISBN==isbnd]
	print(fila_isbn)
	en_la_base.append(isbnd)
	global num_tit
	global num_aut
	global num_edit
	global num_pre
	global num_tem
	try:
		tit = fila_isbn.iat[0,int(num_tit)]
		aut = fila_isbn.iat[0,int(num_aut)]
		edit = fila_isbn.iat[0,int(num_edit)]
		pre = fila_isbn.iat[0,int(num_pre)]
		tem = fila_isbn.iat[0,int(num_tem)]
		en_la_base.append(isbnd)	
		funerariat(tit, isbnd)
		funerariaa(aut, isbnd)
		funerariae(edit, isbnd)
		funerariap(pre, isbnd)
		funerariatema(isbnd, tem)
		concatenadopub(isbnd)
		cuadro_resultado.insert(
			END, dic_titulopub[isbnd] + "," +
			isbnd + "," + imagenes[isbnd] + "," +
			isbnd + "," + "1" + "," + dic_precio[isbnd] + "," + "Nuevo" + "," + "des" + "," + " " + "," +
			"Clásica" + "," + "Mercado Envíos | Mercado Envíos Flex" + "," + "A cargo del comprador" + "," +
			"Acepto" + "," + "Garantía del vendedor"  + "," + "1" + "," + "meses"  + "," + "Papel" + "," + 
			dic_tema[isbnd]  + "," + dic_titulo[isbnd] + "," + dic_autor[isbnd] + "," + "Español" + "," + 
			dic_editorial[isbnd] + "," + dic_tema[isbnd] + "\n"
			)
	except IndexError:
		no_en_la_base.append(isbnd)			
def concatenado(isbnd):
	"""concatena las imágenes de portada y contraportada"""
	if isbnd in portada:
		if isbnd in contraportada:
			if isbnd in paginaextra1:
				if isbnd in paginaextra2:
					if isbnd in paginaextra3:
						if isbnd in paginaextra4:
							imagenes[isbnd] = portada[isbnd] + "; " + contraportada[isbnd] + "; " + paginaextra1[isbnd] + "; " + paginaextra2[isbnd] + "; " + paginaextra3[isbnd] + "; " + paginaextra4[isbnd] + "; https://i.postimg.cc/B6SMfSSh/001.jpg"
						else:
							imagenes[isbnd] = portada[isbnd] + "; " + contraportada[isbnd] + "; " + paginaextra1[isbnd] + "; " + paginaextra2[isbnd] + "; " + paginaextra3[isbnd] + "; https://i.postimg.cc/B6SMfSSh/001.jpg"
					else:
						imagenes[isbnd] = portada[isbnd] + "; " + contraportada[isbnd] + "; " + paginaextra1[isbnd] + "; " + paginaextra2[isbnd] + "; https://i.postimg.cc/B6SMfSSh/001.jpg"
				else:
					imagenes[isbnd] = portada[isbnd] + "; " + contraportada[isbnd] + "; " + paginaextra1[isbnd] + "; https://i.postimg.cc/B6SMfSSh/001.jpg"
			else:
				imagenes[isbnd] = portada[isbnd] + "; " + contraportada[isbnd] + "; https://i.postimg.cc/B6SMfSSh/001.jpg"
				#cuadro_resultado.insert(END, imagenes)
		else:
			imagenes[isbnd] = portada[isbnd] + "; https://i.postimg.cc/B6SMfSSh/001.jpg"
	else:
		imagenes[isbnd] =  "https://i.postimg.cc/B6SMfSSh/001.jpg"

def deconstruirisbns(presibn, listas):
	if presibn[-7] != "0":
		isbn_con_error(presibn)
		#cuadro_resultado.insert(END, "\n" + presibn + " será excluido de la lista final porque no tiene el formato adecuado.\n Recordá que después del ISBN debe incluir 001.jpg\n así se determina la posición de la foto.")
	else:
		if presibn[-20] == "9":
			largo = len(presibn)
			desde = int(largo)-20
			preisbna = presibn[desde:largo]
			isbn = preisbna[0:13]
			resto = preisbna[13:20]
			isbns.append(isbn)
			if resto == "001.jpg":
				portada[isbn] = listas[0]
				listas.pop(0)
			elif resto == "002.jpg":
				contraportada[isbn] = listas[0]
				listas.pop(0)
			elif resto == "003.jpg":
				paginaextra1[isbn] = listas[0]
				listas.pop(0)
			elif resto == "004.jpg":
				paginaextra2[isbn] = listas[0]
				listas.pop(0)
			elif resto == "005.jpg":
				paginaextra3[isbn] = listas[0]
				listas.pop(0)
			elif resto == "006.jpg":
				paginaextra4[isbn] = listas[0]
				listas.pop(0)
		else:
			largo = len(presibn)
			desde = int(largo)-17
			preisbna = presibn[desde:largo]
			isbn = preisbna[0:10]
			resto = preisbna[10:17]
			isbns.append(isbn)
			if resto == "001.jpg":
				portada[isbn] = listas[0]
				listas.pop(0)
			elif resto == "002.jpg":
				contraportada[isbn] = listas[0]
				listas.pop(0)
			elif resto == "003.jpg":
				paginaextra1[isbn] = listas[0]
				listas.pop(0)
			elif resto == "004.jpg":
				paginaextra2[isbn] = listas[0]
				listas.pop(0)
			elif resto == "005.jpg":
				paginaextra3[isbn] = listas[0]
				listas.pop(0)
			elif resto == "006.jpg":
				paginaextra4[isbn] = listas[0]
				listas.pop(0)
def fin():
	mb.showinfo('Aviso', 'Proceso concluido')
def noenlabase():
	for n in no_en_la_base:
		mb.showinfo('Aviso', n + '\n no se encontraban en la base de datos')	

def cargarexcel():
	global excel_concatenador_libro
	libro_ex_conca=excel_concatenador_libro.get()
	global excel_concatenador_hoja
	hoja_ex_conca=excel_concatenador_hoja.get()
	wb = op.load_workbook(libro_ex_conca+'.xlsx')
	global sheet
	sheet = wb[hoja_ex_conca]
	global excel
	excel = True
	toplevel_desdeExcel_concatenador.destroy()
def desdeExcel_concatenador():
	#Toplevel de ingreso de datos
	global toplevel_desdeExcel_concatenador
	toplevel_desdeExcel_concatenador = Toplevel(window)
	toplevel_desdeExcel_concatenador.title("Ingresar Catálogo desde Excel")
	frame_DE = Frame(toplevel_desdeExcel_concatenador)
	frame_DE.pack()
	frame_DE2=Frame(toplevel_desdeExcel_concatenador)
	frame_DE2.pack()
	#Ingresar lista de publicaciones
	excel_concatenador_libro_lab = Label(frame_DE, text="Inserte Libro de publicaciones:")
	excel_concatenador_libro_lab.grid(column=1, row=2)
	excel_concatenador_libro_lab2 = Label(frame_DE, text=".xlsx")
	excel_concatenador_libro_lab2.grid(column=3, row=2)
	global excel_concatenador_libro
	excel_concatenador_libro = Entry(frame_DE, width=15)
	excel_concatenador_libro.grid(column=2, row=2, padx=5, pady=5)
	excel_concatenador_libro.insert(END,"EML")
	excel_concatenador_hoja_lab = Label(frame_DE, text="Inserte nombre de la hoja:")
	excel_concatenador_hoja_lab.grid(column=1, row=3)
	global excel_concatenador_hoja
	excel_concatenador_hoja = Entry(frame_DE, width=15)
	excel_concatenador_hoja.grid(column=2, row=3, padx=5, pady=5)
	excel_concatenador_hoja.insert(END,"EML")
	excel_concatenador_columna_titulo_lab= Label(frame_DE, text="Inserte el número de columna de Título")
	excel_concatenador_columna_titulo_lab.grid(column=1, row=4)
	global excel_concatenador_columna_titulo_ent
	excel_concatenador_columna_titulo_ent= Entry(frame_DE, width=5)
	excel_concatenador_columna_titulo_ent.grid(column=2, row=4, padx=5, pady=5)
	excel_concatenador_columna_titulo_ent.insert(END,"2")
	excel_concatenador_columna_autor_lab= Label(frame_DE, text="Inserte el número de columna de Autor:")
	excel_concatenador_columna_autor_lab.grid(column=1, row=5)
	global excel_concatenador_columna_autor_ent
	excel_concatenador_columna_autor_ent= Entry(frame_DE, width=5)
	excel_concatenador_columna_autor_ent.grid(column=2, row=5, padx=5, pady=5)
	excel_concatenador_columna_autor_ent.insert(END,"3")
	excel_concatenador_columna_editorial_lab= Label(frame_DE, text="Inserte el número de columna de Editorial:")
	excel_concatenador_columna_editorial_lab.grid(column=1, row=6)
	global excel_concatenador_columna_editorial_ent
	excel_concatenador_columna_editorial_ent= Entry(frame_DE, width=5)
	excel_concatenador_columna_editorial_ent.grid(column=2, row=6, padx=5, pady=5)
	excel_concatenador_columna_editorial_ent.insert(END,"4")
	excel_concatenador_columna_isbn_lab= Label(frame_DE, text="Inserte el número de columna de ISBN:")
	excel_concatenador_columna_isbn_lab.grid(column=1, row=7)
	global excel_concatenador_columna_isbn_ent
	excel_concatenador_columna_isbn_ent= Entry(frame_DE, width=5)
	excel_concatenador_columna_isbn_ent.grid(column=2, row=7, padx=5, pady=5)
	excel_concatenador_columna_isbn_ent.insert(END,"9")
	excel_concatenador_columna_precio_lab= Label(frame_DE, text="Inserte el número de columna de Precio:")
	excel_concatenador_columna_precio_lab.grid(column=1, row=8)
	global excel_concatenador_columna_precio_ent
	excel_concatenador_columna_precio_ent= Entry(frame_DE, width=5)
	excel_concatenador_columna_precio_ent.grid(column=2, row=8, padx=5, pady=5)
	excel_concatenador_columna_precio_ent.insert(END,"29")
	excel_concatenador_columna_tema_lab= Label(frame_DE, text="Inserte el número de columna de Tema:")
	excel_concatenador_columna_tema_lab.grid(column=1, row=9)
	global excel_concatenador_columna_tema_ent
	excel_concatenador_columna_tema_ent= Entry(frame_DE, width=5)
	excel_concatenador_columna_tema_ent.grid(column=2, row=9, padx=5, pady=5)
	excel_concatenador_columna_tema_ent.insert(END,"5")
	bot_guardar_catalogo = Button(frame_DE2, text="Guardar", command=cargarexcel)
	bot_guardar_catalogo.pack()
def cargarcsv():
	global Csv_concatenador_libro
	hoja_csv=Csv_concatenador_libro.get()
	global csv
	csv=pd.read_csv(hoja_csv+'.txt',sep=';')
	global num_tit
	num_tit = Csv_concatenador_columna_titulo_ent.get()
	global num_aut
	num_aut = Csv_concatenador_columna_autor_ent.get()
	global num_edit
	num_edit = Csv_concatenador_columna_editorial_ent.get()
	global num_pre
	num_pre = Csv_concatenador_columna_precio_ent.get()
	global num_tem
	num_tem = Csv_concatenador_columna_tema_ent.get()
	global csvcon
	csvcon = True
	toplevel_desdeCsv_concatenador.destroy()
def desdecsv_concatenador():
	#Toplevel de ingreso de datos
	global toplevel_desdeCsv_concatenador
	toplevel_desdeCsv_concatenador = Toplevel(window)
	toplevel_desdeCsv_concatenador.title("Ingresar Catálogo desde Csv")
	frame_DE = Frame(toplevel_desdeCsv_concatenador)
	frame_DE.pack()
	frame_DE2=Frame(toplevel_desdeCsv_concatenador)
	frame_DE2.pack()
	#Ingresar lista de publicaciones
	Csv_concatenador_libro_lab = Label(frame_DE, text="Inserte Libro de publicaciones:")
	Csv_concatenador_libro_lab.grid(column=1, row=2)
	Csv_concatenador_libro_lab2 = Label(frame_DE, text=".csv")
	Csv_concatenador_libro_lab2.grid(column=3, row=2)
	global Csv_concatenador_libro
	Csv_concatenador_libro = Entry(frame_DE, width=15)
	Csv_concatenador_libro.grid(column=2, row=2, padx=5, pady=5)
	Csv_concatenador_libro.insert(END,"EML")
	Csv_concatenador_columna_titulo_lab= Label(frame_DE, text="Inserte el número de columna de Título")
	Csv_concatenador_columna_titulo_lab.grid(column=1, row=4)
	global Csv_concatenador_columna_titulo_ent
	Csv_concatenador_columna_titulo_ent= Entry(frame_DE, width=5)
	Csv_concatenador_columna_titulo_ent.grid(column=2, row=4, padx=5, pady=5)
	Csv_concatenador_columna_titulo_ent.insert(END,"1")
	Csv_concatenador_columna_autor_lab= Label(frame_DE, text="Inserte el número de columna de Autor:")
	Csv_concatenador_columna_autor_lab.grid(column=1, row=5)
	global Csv_concatenador_columna_autor_ent
	Csv_concatenador_columna_autor_ent= Entry(frame_DE, width=5)
	Csv_concatenador_columna_autor_ent.grid(column=2, row=5, padx=5, pady=5)
	Csv_concatenador_columna_autor_ent.insert(END,"2")
	Csv_concatenador_columna_editorial_lab= Label(frame_DE, text="Inserte el número de columna de Editorial:")
	Csv_concatenador_columna_editorial_lab.grid(column=1, row=6)
	global Csv_concatenador_columna_editorial_ent
	Csv_concatenador_columna_editorial_ent= Entry(frame_DE, width=5)
	Csv_concatenador_columna_editorial_ent.grid(column=2, row=6, padx=5, pady=5)
	Csv_concatenador_columna_editorial_ent.insert(END,"3")
	Csv_concatenador_columna_isbn_lab= Label(frame_DE, text="Inserte el número de columna de ISBN:")
	Csv_concatenador_columna_isbn_lab.grid(column=1, row=7)
	global Csv_concatenador_columna_isbn_ent
	Csv_concatenador_columna_isbn_ent= Entry(frame_DE, width=5)
	Csv_concatenador_columna_isbn_ent.grid(column=2, row=7, padx=5, pady=5)
	Csv_concatenador_columna_isbn_ent.insert(END,"8")
	Csv_concatenador_columna_precio_lab= Label(frame_DE, text="Inserte el número de columna de Precio:")
	Csv_concatenador_columna_precio_lab.grid(column=1, row=8)
	global Csv_concatenador_columna_precio_ent
	Csv_concatenador_columna_precio_ent= Entry(frame_DE, width=5)
	Csv_concatenador_columna_precio_ent.grid(column=2, row=8, padx=5, pady=5)
	Csv_concatenador_columna_precio_ent.insert(END,"28")
	Csv_concatenador_columna_tema_lab= Label(frame_DE, text="Inserte el número de columna de Tema:")
	Csv_concatenador_columna_tema_lab.grid(column=1, row=9)
	global Csv_concatenador_columna_tema_ent
	Csv_concatenador_columna_tema_ent= Entry(frame_DE, width=5)
	Csv_concatenador_columna_tema_ent.grid(column=2, row=9, padx=5, pady=5)
	Csv_concatenador_columna_tema_ent.insert(END,"4")
	bot_guardar_catalogo = Button(frame_DE2, text="Guardar", command=cargarcsv)
	bot_guardar_catalogo.pack()

def concatenar():
	global excel
	global csvcon
	if excel == False and csvcon == False:
		mb.showerror("Error","No se ha cargado el catálogo")
	else:		
		lista = ingreso.get()
		listas = lista.split()
		preisbns = listas[:]
		while len(listas)>0:
			for d in preisbns:
				deconstruirisbns(d, listas)
			for i in isbns:
				if i not in isbn_depurado:
					isbn_depurado.append(i)
			for isbnd in isbn_depurado:
				concatenado(isbnd)
				print(isbnd)
				
				if excel==True:
					global sheet
					ga = (len(sheet['A'])+1)
					for x in range (1,ga):
						xrow = sheet.cell(row=x, column=8).value
						busqueda(xrow, isbnd, x)		
				elif csvcon==True:
					busqudacsv(isbnd) 		
				#cuadro_resultado.insert(END, dic_titulopub[isbnd] + "," + isbnd + "," + imagenes[isbnd] + "," + isbnd + "," + dic_titulo[isbnd] + "," + dic_autor[isbnd] + "," + dic_editorial[isbnd])
		noenlabase()
		fin()
def isbn_con_error(isbn):
	mb.showinfo('Aviso', isbn + ' no pudo ser concatenado por no tener el formato adecuado')



def formato_correcto():
	mb.showinfo(
		'Formato correcto',
		'Los URLs deben ser ingresados en una línea separados por espacios.'
		+ '\n con la forma: http://***/ISBN001.jpg \n' +
		'Donde 001.jpg será la portada, 002.jpg será la contraportada \n'
		+ 'se puede incluir hasta 006.jpg.\n' +
		'El concatenador soporta ISBN 10 y EAN13'
		)


#GUI

from tkinter import*
from tkinter import messagebox as mb

def excluidos():
	nv = Toplevel(window)
	cuadro_excluidos = Text(nv, width=75, height=25)
	cuadro_excluidos.pack()
	for isbndn in no_en_la_base:
		cuadro_excluidos.insert(END, isbndn + "," + imagenes[isbndn] + "\n") 

window = Tk()
window.title("Librería Losada")


frame = Frame(window, width=1200, height=500)
frame.pack()

bienvenida = Label(frame, text="Concatenador de imágenes")
bienvenida.grid(column=1, row=0)


ingrese = Label(
	frame,
	text="Ingrese URL de imágenes:"
	)

ingrese.grid(column=0, row=1)

ingreso = Entry(frame, width=75)
ingreso.grid(column=1, row=1, padx=5, pady=5)

boton = Button(frame, text="Concatenar URL", width=15, height=5, command=concatenar)
boton.grid(column=3, row=3,padx=5, pady=5)

boton_excel_concatenador = Button(frame, text="Cargar Excel", width=15, height=5, command=desdeExcel_concatenador)
boton_excel_concatenador.grid(column=4, row=5,padx=5, pady=5)
boton_csv_concatenador = Button(frame, text="Cargar CSV", width=15, height=5, command=desdecsv_concatenador)
boton_csv_concatenador.grid(column=5, row=5,padx=5, pady=5)

resultado = Label(frame, text="URL concatenados:")
resultado.grid(column=0, row=2, padx=5, pady=5),

cuadro_resultado = Text(frame, width=75, height=25)
cuadro_resultado.grid(column=1, row=3, padx=5, pady=5)

scroll = Scrollbar(frame, command=cuadro_resultado.yview)
scroll.grid(column=2, row=3, sticky="nsew")
cuadro_resultado.config(yscrollcommand=scroll.set)

boton_de_formato = Button(frame, text="Ver formato correcto", command=formato_correcto)
boton_de_formato.grid(column=3, row=1,padx=5, pady=5)

boton_de_formato = Button(frame, text="Ver URLs excluidos", command=excluidos)
boton_de_formato.grid(column=3, row=4,padx=5, pady=5)

window.mainloop()
