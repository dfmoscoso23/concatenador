#Backstage
import openpyxl as op
wb = op.load_workbook('ML.xlsx')
sheet = wb['ML']
#Diccionarios
dic_titulo = {}
dic_autor = {}
dic_editorial = {}
dic_precio = {}
dic_titulopub = {}
portada = {}
contraportada = {}
paginaextra1 = {}
paginaextra2 = {}
paginaextra3 = {}
paginaextra4 = {}
dic_tema = {}
apellido = {}
imagenes = {}
isbns = []
isbn_depurado = []
en_la_base = []
no_en_la_base = []

def funerariat(tit, isbnd):
	if ',' in tit:
		titu = tit.split(',')
		arti = titu [1].strip(" ")
		titul = titu[0].strip(" ")
		titulo = arti + " " + titul
	else:
		titulo = tit
	tituloo = titulo.title()
	titulod = tituloo.strip()
	dic_titulo[isbnd] = titulod

def funerariaa(aut, isbnd):
	if ',' in aut:
		auto = aut.split(',')
		artic = auto [1].strip(" ")
		ape = auto[0].strip(" ")
		autor = artic + " " + ape
		apellido[isbnd] = ape.title()
	else:
		autor = aut.strip("	")
		apellido[isbnd] = aut.title()
	autort = autor.strip(" ")
	autord = autort.title()
	dic_autor[isbnd] = autord


def funerariae(edit, isbnd):
	editora = edit.title()
	editorial = editora.strip(" ")
	dic_editorial[isbnd] = editorial
def funerariap(pre, isbnd):
	dic_precio[isbnd] =str(pre)	

def concatenadopub(isbnd):
	dic_titulopub[isbnd] = dic_titulo[isbnd] + " - "+ apellido[isbnd] + " - " + dic_editorial[isbnd]

def funerariatema(isbnd, tem):
	dic_tema[isbnd] = tem.title()

def busqueda(xrow, isbnd, x):
	#cuadro_resultado.insert(END, isbnd)
	#cuadro_resultado.insert(END, xrow)
	if isbnd==str(xrow):
		en_la_base.append(isbnd)	
		#print(sheet.cell(row=x, column=2).value)
		tit = sheet.cell(row=x, column=2).value
		aut = sheet.cell(row=x, column=3).value
		edit = sheet.cell(row=x, column=4).value
		pre = sheet.cell(row=x, column=11).value
		tem = sheet.cell(row=x, column=6).value
		funerariat(tit, isbnd)
		funerariaa(aut, isbnd)
		funerariae(edit, isbnd)
		funerariap(pre, isbnd)
		funerariatema(isbnd, tem)
		concatenadopub(isbnd)
		cuadro_resultado.insert(
			END, dic_titulopub[isbnd] + "," +
			isbnd + "," + imagenes[isbnd] + "," +
			isbnd + "," + "1" + "," + dic_precio[isbnd] + "," + "Nuevo" + "," + "des" + "," + " " + "," +
			"Clásica" + "," + "Mercado Envíos | Mercado Envíos Flex" + "," + "A cargo del comprador" + "," +
			"Acepto" + "," + "Garantía del vendedor"  + "," + "1" + "," + "meses"  + "," + "Papel" + "," + 
			dic_tema[isbnd]  + "," + dic_titulo[isbnd] + "," + dic_autor[isbnd] + "," + "Español" + "," + 
			dic_editorial[isbnd] + "," + dic_tema[isbnd] + "\n"
			)
def concatenado(isbnd):
	"""concatena las imágenes de portada y contraportada"""
	if isbnd in portada:
		if isbnd in contraportada:
			if isbnd in paginaextra1:
				if isbnd in paginaextra2:
					if isbnd in paginaextra3:
						if isbnd in paginaextra4:
							imagenes[isbnd] = portada[isbnd] + "; " + contraportada[isbnd] + "; " + paginaextra1[isbnd] + "; " + paginaextra2[isbnd] + "; " + paginaextra3[isbnd] + "; " + paginaextra4[isbnd] + "; https://i.postimg.cc/B6SMfSSh/001.jpg"
						else:
							imagenes[isbnd] = portada[isbnd] + "; " + contraportada[isbnd] + "; " + paginaextra1[isbnd] + "; " + paginaextra2[isbnd] + "; " + paginaextra3[isbnd] + "; https://i.postimg.cc/B6SMfSSh/001.jpg"
					else:
						imagenes[isbnd] = portada[isbnd] + "; " + contraportada[isbnd] + "; " + paginaextra1[isbnd] + "; " + paginaextra2[isbnd] + "; https://i.postimg.cc/B6SMfSSh/001.jpg"
				else:
					imagenes[isbnd] = portada[isbnd] + "; " + contraportada[isbnd] + "; " + paginaextra1[isbnd] + "; https://i.postimg.cc/B6SMfSSh/001.jpg"
			else:
				imagenes[isbnd] = portada[isbnd] + "; " + contraportada[isbnd] + "; https://i.postimg.cc/B6SMfSSh/001.jpg"
				#cuadro_resultado.insert(END, imagenes)
		else:
			imagenes[isbnd] = portada[isbnd] + "; https://i.postimg.cc/B6SMfSSh/001.jpg"
	else:
		imagenes[isbnd] =  "https://i.postimg.cc/B6SMfSSh/001.jpg"

def deconstruirisbns(presibn, listas):
	if presibn[-7] != "0":
		isbn_con_error(presibn)
		#cuadro_resultado.insert(END, "\n" + presibn + " será excluido de la lista final porque no tiene el formato adecuado.\n Recordá que después del ISBN debe incluir 001.jpg\n así se determina la posición de la foto.")
	else:
		if presibn[-20] == "9":
			largo = len(presibn)
			desde = int(largo)-20
			preisbna = presibn[desde:largo]
			isbn = preisbna[0:13]
			resto = preisbna[13:20]
			isbns.append(isbn)
			if resto == "001.jpg":
				portada[isbn] = listas[0]
				listas.pop(0)
			elif resto == "002.jpg":
				contraportada[isbn] = listas[0]
				listas.pop(0)
			elif resto == "003.jpg":
				paginaextra1[isbn] = listas[0]
				listas.pop(0)
			elif resto == "004.jpg":
				paginaextra2[isbn] = listas[0]
				listas.pop(0)
			elif resto == "005.jpg":
				paginaextra3[isbn] = listas[0]
				listas.pop(0)
			elif resto == "006.jpg":
				paginaextra4[isbn] = listas[0]
				listas.pop(0)
		else:
			largo = len(presibn)
			desde = int(largo)-17
			preisbna = presibn[desde:largo]
			isbn = preisbna[0:10]
			resto = preisbna[10:17]
			isbns.append(isbn)
			if resto == "001.jpg":
				portada[isbn] = listas[0]
				listas.pop(0)
			elif resto == "002.jpg":
				contraportada[isbn] = listas[0]
				listas.pop(0)
			elif resto == "003.jpg":
				paginaextra1[isbn] = listas[0]
				listas.pop(0)
			elif resto == "004.jpg":
				paginaextra2[isbn] = listas[0]
				listas.pop(0)
			elif resto == "005.jpg":
				paginaextra3[isbn] = listas[0]
				listas.pop(0)
			elif resto == "006.jpg":
				paginaextra4[isbn] = listas[0]
				listas.pop(0)
def fin():
	mb.showinfo('Aviso', 'Proceso concluido')
def noenlabase():
	for n in no_en_la_base:
		mb.showinfo('Aviso', n + '\n no se encontraban en la base de datos')	
def concatenar():
	lista = ingreso.get()
	listas = lista.split()
	preisbns = listas[:]
	while len(listas)>1:
		for d in preisbns:
			deconstruirisbns(d, listas)
		for i in isbns:
			if i not in isbn_depurado:
				isbn_depurado.append(i)
		for isbnd in isbn_depurado:
			concatenado(isbnd)
			ga = (len(sheet['A'])+1)
			for x in range (1,ga):
				xrow = sheet.cell(row=x, column=8).value
				busqueda(xrow, isbnd, x)
			#cuadro_resultado.insert(END, dic_titulopub[isbnd] + "," + isbnd + "," + imagenes[isbnd] + "," + isbnd + "," + dic_titulo[isbnd] + "," + dic_autor[isbnd] + "," + dic_editorial[isbnd])
	for isbnd in isbn_depurado:
		if isbnd not in en_la_base:
			no_en_la_base.append(isbnd)
	noenlabase()
	fin()
def isbn_con_error(isbn):
	mb.showinfo('Aviso', isbn + ' no pudo ser concatenado por no tener el formato adecuado')



def formato_correcto():
	mb.showinfo(
		'Formato correcto',
		'Los URLs deben ser ingresados en una línea separados por espacios.'
		+ '\n con la forma: http://***/ISBN001.jpg \n' +
		'Donde 001.jpg será la portada, 002.jpg será la contraportada \n'
		+ 'se puede incluir hasta 006.jpg.\n' +
		'El concatenador soporta ISBN 10 y EAN13'
		)


#GUI

from tkinter import*
from tkinter import messagebox as mb

def excluidos():
	nv = Toplevel(window)
	cuadro_excluidos = Text(nv, width=75, height=25)
	cuadro_excluidos.pack()
	for isbndn in no_en_la_base:
		cuadro_excluidos.insert(END, isbndn + "," + imagenes[isbndn] + "\n") 

window = Tk()
window.title("Librería Losada")


frame = Frame(window, width=1200, height=500)
frame.pack()

bienvenida = Label(frame, text="Concatenador de imágenes")
bienvenida.grid(column=1, row=0)


ingrese = Label(
	frame,
	text="Ingrese URL de imágenes:"
	)

ingrese.grid(column=0, row=1)

ingreso = Entry(frame, width=75)
ingreso.grid(column=1, row=1, padx=5, pady=5)

boton = Button(frame, text="Concatenar URL", width=15, height=5, command=concatenar)
boton.grid(column=3, row=3,padx=5, pady=5)

resultado = Label(frame, text="URL concatenados:")
resultado.grid(column=0, row=2, padx=5, pady=5),

cuadro_resultado = Text(frame, width=75, height=25)
cuadro_resultado.grid(column=1, row=3, padx=5, pady=5)

scroll = Scrollbar(frame, command=cuadro_resultado.yview)
scroll.grid(column=2, row=3, sticky="nsew")
cuadro_resultado.config(yscrollcommand=scroll.set)

boton_de_formato = Button(frame, text="Ver formato correcto", command=formato_correcto)
boton_de_formato.grid(column=3, row=1,padx=5, pady=5)

boton_de_formato = Button(frame, text="Ver URLs excluidos", command=excluidos)
boton_de_formato.grid(column=3, row=4,padx=5, pady=5)

window.mainloop()
